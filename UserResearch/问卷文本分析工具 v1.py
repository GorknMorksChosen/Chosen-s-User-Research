import streamlit as st
import pandas as pd
from survey_tools.utils.io import read_table_auto
from survey_tools.utils.download_filename import safe_download_filename
from survey_tools.utils.wjx_header import normalize_wjx_headers
import os
import json
import plotly.express as px
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
import io
import time
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from collections import Counter
import jieba
import jieba.analyse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pydantic import BaseModel, Field
from survey_core_quant import make_safe_sheet_name
from survey_tools.core.question_type import (
    detect_column_type,
    get_prefix,
    infer_type_from_columns,
    parse_columns_for_questions,
)
from survey_tools.core.quant import extract_qnum
from survey_tools.core.survey_metadata_columns import is_metadata_column
from survey_tools.web.outline_upload import (
    OUTLINE_PLATFORM_OPTIONS,
    parse_uploaded_outline_file,
)

# --- 1. 配置与初始化 ---
st.set_page_config(page_title="通用问卷文本分析引擎", layout="wide")

class UniversalAnalysis(BaseModel):
    item_id: str = Field(description="必须原样返回输入数据中该条样本对应的 id（字符串格式）")
    player_segment: str = Field(description="玩家分群标签。必须严格、完整地继承输入数据中提供的 segment 字段内容。如果输入中 segment 有值，严禁进行任何翻译、改写、添加后缀或修饰。只有当 segment 为空或为 None 时，才允许根据文本简要推断分群。")
    key_needs: str = Field(description="围绕研究目标提炼的核心诉求关键词，使用“/”分隔多个短语")
    qualitative_insight: str = Field(description="围绕【研究目标】给出的定性洞察总结，可以概括玩家的典型体验维度、画像特征或动机结构")
    emotion_label: str = Field(description="情绪倾向标签，建议取值：正面 / 中性 / 负面")
    emotion_score: int = Field(description="情感强度评分，1-10 的整数，数值越高情绪越强烈")


def extract_json_str(text: str) -> str:
    if not isinstance(text, str):
        return ""
    s = text.strip()
    if "```" in s:
        parts = s.split("```")
        candidate = max(parts, key=len)
        s = candidate.strip()
    start_arr = s.find("[")
    end_arr = s.rfind("]")
    if start_arr != -1 and end_arr != -1 and end_arr > start_arr:
        return s[start_arr : end_arr + 1].strip()
    start_obj = s.find("{")
    end_obj = s.rfind("}")
    if start_obj != -1 and end_obj != -1 and end_obj > start_obj:
        return s[start_obj : end_obj + 1].strip()
    return s


ERROR_PLACEHOLDER = {
    "player_segment": "分析失败",
    "key_needs": "API异常",
    "qualitative_insight": "分析超时或被拦截",
    "emotion_label": "中性",
    "emotion_score": 0,
}


def safe_json_parse(text: str):
    if not isinstance(text, str):
        raise ValueError("LLM 返回内容不是字符串")
    s = text
    s = re.sub(r"```[\w]*", "", s)
    s = s.replace("```", "")
    json_str = extract_json_str(s)
    return json.loads(json_str)


def invoke_llm_with_retry(messages, model_name, api_key, base_url, temperature=0.3, timeout=45, max_retries=2, backoff_base=1.0, backoff_max=8.0):
    last_error = None
    for attempt in range(max_retries + 1):
        try:
            llm = ChatOpenAI(
                model=model_name,
                openai_api_key=api_key,
                openai_api_base=base_url,
                temperature=temperature,
                max_retries=0,
                timeout=timeout,
            )
            response = llm.invoke(messages)
            return response, attempt
        except Exception as e:
            last_error = e
            if attempt >= max_retries:
                break
            sleep_s = min(backoff_max, backoff_base * (2 ** attempt))
            sleep_s += random.uniform(0, backoff_base)
            time.sleep(sleep_s)
    raise last_error


def detect_spoiler_columns(target_cols, context_cols, api_key, base_url, model_name, max_retries=2, backoff_base=1.0, backoff_max=8.0):
    if not target_cols or not context_cols or not api_key:
        return []
    try:
        tmpl = ChatPromptTemplate.from_template(
            """
你是一名严谨的问卷设计和语言专家，负责审查“背景参考列”是否会对文本填空题造成“剧透”。

你会得到两个数组：
- target_columns: 需要分析的开放式填空题标题
- context_columns: 背景参考列标题

请判断每一个 context 列是否本质上是 target 列的结构化版本，或者直接泄露了 target 列的答案。例如：
- target: "Q18. 请描述你未来最期待的英雄角色" 
  context: "Q16. 期望英雄类型（可多选）"  → 这是剧透 (Spoiler)

输出要求（非常重要）：
1. 只输出一个 JSON 对象。
2. 结构必须为：
   {{
     "spoiler_context_columns": ["列名1", "列名2", ...]
   }}
3. 只填写你认为存在明显语义重叠、会影响“盲测”纯净度的背景列；没有则返回空数组。

target_columns:
{target_cols_json}

context_columns:
{context_cols_json}
            """
        )
        messages = tmpl.format_messages(
            target_cols_json=json.dumps(target_cols, ensure_ascii=False),
            context_cols_json=json.dumps(context_cols, ensure_ascii=False),
        )
        resp, _ = invoke_llm_with_retry(
            messages,
            model_name=model_name,
            api_key=api_key,
            base_url=base_url,
            temperature=0.0,
            timeout=20,
            max_retries=max_retries,
            backoff_base=backoff_base,
            backoff_max=backoff_max,
        )
        data = safe_json_parse(resp.content)
        if isinstance(data, dict):
            arr = data.get("spoiler_context_columns", [])
        elif isinstance(data, list):
            arr = data
        else:
            return []
        spoilers = [c for c in arr if c in context_cols]
        return spoilers
    except Exception:
        return []


def sanitize_value(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if not s:
        return None
    low = s.lower()
    if low in ("none", "nan"):
        return None
    if s in ("否", "无"):
        return None
    return s


def build_context_string(row, context_cols):
    parts = []
    for c in context_cols:
        if c in row:
            val = sanitize_value(row[c])
            if val is not None:
                parts.append(f"{c}: {val}")
    return " | ".join(parts)


def build_target_string(row, target_cols):
    parts = []
    for t in target_cols:
        if t in row:
            val = sanitize_value(row[t])
            if val is not None:
                parts.append(f"{t}: {val}")
    return " | ".join(parts)


def build_text_tool_column_type_map(df: pd.DataFrame, outline: dict | None = None) -> dict[str, str]:
    """为文本工具构建列题型映射（与 Quant/Pipeline 口径尽量对齐）。"""
    columns = df.columns.tolist()
    questions_data = parse_columns_for_questions(columns)

    q_num_to_type: dict[int, str] = {}
    for q_num, info in questions_data.items():
        raw = infer_type_from_columns(info)
        if raw:
            if "矩阵" in raw:
                q_num_to_type[q_num] = "矩阵"
            elif "多选" in raw:
                q_num_to_type[q_num] = "多选"
            elif "NPS" in raw:
                q_num_to_type[q_num] = "NPS"
            elif "评分" in raw:
                q_num_to_type[q_num] = "评分"
            elif "填空" in raw or "开放" in raw:
                q_num_to_type[q_num] = "开放文本"
            else:
                q_num_to_type[q_num] = "单选"

    if outline:
        for q_num, info in outline.items():
            otype = str(info.get("type", ""))
            if "矩阵" in otype:
                if "文本" in otype or "填空" in otype:
                    q_num_to_type[q_num] = "开放文本"
                else:
                    q_num_to_type[q_num] = "矩阵"
            elif "多选" in otype:
                q_num_to_type[q_num] = "多选"
            elif "NPS" in otype.upper() or "nps" in otype.lower():
                q_num_to_type[q_num] = "NPS"
            elif "填空" in otype or "文本" in otype:
                q_num_to_type[q_num] = "开放文本"
            elif "量表" in otype or "评分" in otype:
                q_num_to_type[q_num] = "评分"
            elif "单选" in otype:
                q_num_to_type[q_num] = "单选"

    known_multi_prefixes: set[str] = set()
    for col in columns:
        q_str = extract_qnum(str(col))
        if q_str:
            try:
                if q_num_to_type.get(int(q_str)) == "多选":
                    known_multi_prefixes.add(get_prefix(str(col)))
            except ValueError:
                pass

    type_map: dict[str, str] = {}
    for col in columns:
        col_s = str(col)
        if is_metadata_column(col_s):
            type_map[col_s] = "元数据"
            continue
        q_str = extract_qnum(col_s)
        if q_str:
            try:
                q_num = int(q_str)
                if q_num in q_num_to_type:
                    type_map[col_s] = q_num_to_type[q_num]
                    continue
            except ValueError:
                pass
        detected = detect_column_type(col_s, df[col], get_prefix(col_s), known_multi_prefixes)
        type_map[col_s] = "开放文本" if detected == "排序" else detected
    return type_map


def build_question_selector_map(df: pd.DataFrame, col_type_map: dict[str, str]) -> dict[str, list[str]]:
    """构建题目级选择映射：显示标签 -> 该题对应的列集合。"""
    q_data = parse_columns_for_questions(df.columns.tolist())
    label_map: dict[str, list[str]] = {}
    def _norm_qnum_for_cmp(v: str | int | None) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s:
            return ""
        try:
            return str(int(s))
        except ValueError:
            return s

    for q_num in sorted(q_data.keys()):
        info = q_data[q_num]
        q_cols = [
            c
            for c in df.columns
            if _norm_qnum_for_cmp(extract_qnum(str(c))) == _norm_qnum_for_cmp(q_num)
        ]
        if not q_cols:
            continue
        type_candidates = [col_type_map.get(str(c), "未识别") for c in q_cols]
        type_candidates = [t for t in type_candidates if t not in ("元数据", "未识别")]
        q_type = type_candidates[0] if type_candidates else col_type_map.get(str(q_cols[0]), "未识别")
        stem = str(info.get("stem") or q_cols[0]).strip()
        label = f"Q{int(q_num):03d} [{q_type}] {stem}"
        base_label = label
        idx = 2
        while label in label_map:
            label = f"{base_label} ({idx})"
            idx += 1
        label_map[label] = [str(c) for c in q_cols]

    # 无题号列作为“单列题目”补充到题目选择器，避免信息丢失（如 type.玩家分类）
    for c in df.columns:
        c_str = str(c)
        if extract_qnum(c_str):
            continue
        t = col_type_map.get(c_str, "未识别")
        if t == "元数据":
            continue
        label = f"[列:{t}] {c_str}"
        base_label = label
        idx = 2
        while label in label_map:
            label = f"{base_label} ({idx})"
            idx += 1
        label_map[label] = [c_str]
    return label_map


def apply_manual_q_type_overrides(
    df: pd.DataFrame,
    col_type_map: dict[str, str],
    manual_q_type_overrides: dict[str, str] | None,
) -> dict[str, str]:
    """按题号应用手动题型覆盖，供目标/背景题选择器即时使用。"""
    overrides = manual_q_type_overrides or {}
    if not overrides:
        return col_type_map
    updated = dict(col_type_map)

    def _norm_qkey(q: str) -> str:
        q_s = str(q).strip()
        if not q_s:
            return ""
        try:
            return str(int(q_s))
        except ValueError:
            return q_s

    normalized_overrides = {_norm_qkey(k): v for k, v in overrides.items()}
    for col in df.columns:
        col_s = str(col)
        if is_metadata_column(col_s):
            continue
        q_str = extract_qnum(col_s)
        if not q_str:
            continue
        forced_type = normalized_overrides.get(_norm_qkey(q_str))
        if forced_type:
            updated[col_s] = forced_type
    return updated


def compute_keyword_deviations(df, target_cols, group_col, keywords, threshold=0.15, topk=15):
    if df is None or df.empty or not target_cols or not group_col or not keywords:
        return []
    total_n = len(df)
    if total_n == 0:
        return []
    overall_stats = {}
    for kw in keywords:
        try:
            mask = df[target_cols].astype(str).apply(lambda x: x.str.contains(re.escape(kw), na=False)).any(axis=1)
        except Exception:
            mask = pd.Series([False] * total_n)
        overall_mention_count = int(mask.sum())
        overall_stats[kw] = overall_mention_count / total_n if total_n > 0 else 0.0
    deviation_results = []
    grouped = df.groupby(group_col, dropna=False)
    for seg_name, group_df in grouped:
        group_n = len(group_df)
        if group_n == 0:
            continue
        for kw in keywords:
            try:
                g_mask = group_df[target_cols].astype(str).apply(lambda x: x.str.contains(re.escape(kw), na=False)).any(axis=1)
            except Exception:
                g_mask = pd.Series([False] * group_n)
            group_mention_rate = int(g_mask.sum()) / group_n
            overall_rate = overall_stats.get(kw, 0.0)
            deviation = group_mention_rate - overall_rate
            if abs(deviation) >= threshold:
                deviation_results.append(
                    {
                        "segment": str(seg_name),
                        "keyword": kw,
                        "overall_rate": overall_rate,
                        "group_rate": group_mention_rate,
                        "deviation": deviation,
                    }
                )
    if not deviation_results:
        return []
    sorted_results = sorted(deviation_results, key=lambda x: abs(x["deviation"]), reverse=True)
    for r in sorted_results:
        r["overall_rate_str"] = f"{r['overall_rate']:.1%}"
        r["group_rate_str"] = f"{r['group_rate']:.1%}"
        r["deviation_str"] = f"{r['deviation']:+.1%}"
    return sorted_results[:topk]


def prepare_evidence_pool(df, target_cols, group_col, significant_items, max_quotes_per_item=5):
    pool = {}
    if df is None or df.empty or not significant_items:
        return pool
    for item in significant_items:
        kw = item.get("keyword")
        seg = item.get("segment")
        if not kw or seg is None:
            continue
        try:
            mask_seg = df[group_col].astype(str) == str(seg)
        except Exception:
            continue
        try:
            mask_kw = df[target_cols].astype(str).apply(lambda x: x.str.contains(re.escape(kw), na=False)).any(axis=1)
        except Exception:
            mask_kw = pd.Series([False] * len(df))
        mask = mask_seg & mask_kw
        subset = df[mask]
        if subset.empty:
            continue
        subset = subset.head(max_quotes_per_item)
        quotes = []
        for _, row in subset.iterrows():
            parts = []
            for c in target_cols:
                if c in row and pd.notna(row[c]):
                    parts.append(str(row[c]))
            text_snippet = " ".join(parts).strip()
            if not text_snippet:
                continue
            sid = row.get("Internal_ID", None)
            prefix = f"[#ID:{sid}]" if sid is not None else ""
            snippet = text_snippet[:100]
            quotes.append(f"{prefix} {snippet}..." if prefix else f"{snippet}...")
        if quotes:
            key = f"{seg}_{kw}"
            pool[key] = quotes
    return pool


def is_valid_text(text: str, min_length: int) -> bool:
    if text is None:
        return False
    s = str(text).strip()
    if not s:
        return False
    lower = s.lower()
    blacklist = {"不知道", "没有", "无", "略", "没什么", "如题", "111", "暂无"}
    if lower in blacklist:
        return False
    s_no_space = re.sub(r"\s+", "", s)
    if not s_no_space:
        return False
    if len(s_no_space) >= 3:
        for pat_len in (1, 2):
            if len(s_no_space) % pat_len == 0:
                pattern = s_no_space[:pat_len]
                if pattern and pattern * (len(s_no_space) // pat_len) == s_no_space:
                    return False
    core = re.sub(r"[^\w\u4e00-\u9fff]+", "", s_no_space)
    if len(core) < max(1, int(min_length)):
        return False
    return True


def get_keyword_stats(df, columns, top_n: int = 10):
    keyword_stats = {}
    if df is None or not columns:
        return keyword_stats
    for col in columns:
        if col not in df.columns:
            continue
        ser = df[col].dropna().astype(str)
        counter = Counter()
        for text in ser:
            s = text.strip()
            if not s:
                continue
            tags = jieba.analyse.extract_tags(
                s,
                topK=top_n * 2,
                withWeight=False,
                allowPOS=("n", "nr", "ns", "nt", "nz", "v", "vn", "a", "an"),
            )
            for t in tags:
                t = str(t).strip()
                if len(t) < 2:
                    continue
                counter[t] += 1
        if counter:
            keyword_stats[col] = counter.most_common(top_n)
    return keyword_stats


def compute_keyword_group_stats(df, target_cols, group_col, keywords):
    if df is None or not target_cols or not keywords or group_col not in df.columns:
        return pd.DataFrame(), {}
    records = []
    structured = {}
    for col in target_cols:
        if col not in df.columns:
            continue
        ser = df[col].astype(str).fillna("")
        for kw in keywords:
            kw_str = str(kw).strip()
            if not kw_str:
                continue
            mask_all = ser.str.contains(re.escape(kw_str), na=False)
            total_mention_users = int(mask_all.sum())
            if col not in structured:
                structured[col] = {}
            info = {"total_mention_users": total_mention_users, "by_segment": {}}
            grouped = df.groupby(group_col, dropna=False)
            for seg_value, g in grouped:
                g_ser = g[col].astype(str).fillna("")
                mask = g_ser.str.contains(re.escape(kw_str), na=False)
                mention_users = int(mask.sum())
                group_n = int(len(g_ser))
                if group_n == 0:
                    continue
                mention_rate = mention_users / group_n
                info["by_segment"][str(seg_value)] = {
                    "mention_users": mention_users,
                    "mention_rate": mention_rate,
                    "group_size": group_n,
                }
                records.append(
                    {
                        "题目": col,
                        "核心分组": seg_value,
                        "关键词": kw_str,
                        "提及人数": mention_users,
                        "提及率": mention_rate,
                        "组样本数": group_n,
                    }
                )
            structured[col][kw_str] = info
    if records:
        stats_df = pd.DataFrame(records)
    else:
        stats_df = pd.DataFrame(
            columns=["题目", "核心分组", "关键词", "提及人数", "提及率", "组样本数"]
        )
    return stats_df, structured


def compute_keyword_counts_per_question(df, target_cols, keywords):
    result = {}
    if df is None or not target_cols or not keywords:
        return result
    for col in target_cols:
        if col not in df.columns:
            continue
        ser = df[col].astype(str).fillna("")
        col_stats = {}
        for kw in keywords:
            kw_str = str(kw).strip()
            if not kw_str:
                continue
            mask = ser.str.contains(re.escape(kw_str), na=False)
            count_users = int(mask.sum())
            if count_users > 0:
                col_stats[kw_str] = count_users
        if col_stats:
            result[col] = col_stats
    return result

def init_session_state():
    if "df" not in st.session_state:
        st.session_state.df = None
    if "analyzed_df" not in st.session_state:
        st.session_state.analyzed_df = None
    if "api_key" not in st.session_state:
        st.session_state.api_key = ""
    if "base_url" not in st.session_state:
        st.session_state.base_url = "https://api.openai.com/v1"
    if "stop_analysis" not in st.session_state:
        st.session_state.stop_analysis = False
    if "model_name" not in st.session_state:
        st.session_state.model_name = "gpt-3.5-turbo"
    if "batch_size" not in st.session_state:
        st.session_state.batch_size = 3
    if "max_workers" not in st.session_state:
        st.session_state.max_workers = 5
    if "deep_report" not in st.session_state:
        st.session_state.deep_report = ""
    if "segment_report" not in st.session_state:
        st.session_state.segment_report = ""
    if "segment_report_name" not in st.session_state:
        st.session_state.segment_report_name = ""
    if "ppt_outline" not in st.session_state:
        st.session_state.ppt_outline = ""
    if "core_segment_col" not in st.session_state:
        st.session_state.core_segment_col = None
    if "enable_text_filter" not in st.session_state:
        st.session_state.enable_text_filter = True
    if "min_valid_text_length" not in st.session_state:
        st.session_state.min_valid_text_length = 3
    if "segment_reports" not in st.session_state:
        st.session_state.segment_reports = {}
    if "target_cols" not in st.session_state:
        st.session_state.target_cols = []
    if "keyword_stats_global" not in st.session_state:
        st.session_state.keyword_stats_global = {}
    if "keyword_terms_text" not in st.session_state:
        st.session_state.keyword_terms_text = ""
    if "keyword_terms_confirmed" not in st.session_state:
        st.session_state.keyword_terms_confirmed = []
    if "keyword_group_stats_df" not in st.session_state:
        st.session_state.keyword_group_stats_df = None
    if "keyword_struct_for_prompt" not in st.session_state:
        st.session_state.keyword_struct_for_prompt = {}
    if "config_locked" not in st.session_state:
        st.session_state.config_locked = False
    if "locked_config" not in st.session_state:
        st.session_state.locked_config = {}
    if "research_goal" not in st.session_state:
        st.session_state.research_goal = "分析玩家对于新角色/英雄的核心诉求，提炼功能定位、操作特征及外形偏好。"
    if "enable_emotion_analysis" not in st.session_state:
        st.session_state.enable_emotion_analysis = True
    if "user_dict_text" not in st.session_state:
        st.session_state.user_dict_text = "搜打撤, 霸体, 僵直, 弹刀, 韧性条"
    if "user_dict_terms" not in st.session_state:
        st.session_state.user_dict_terms = []
    if "user_dict_applied_hash" not in st.session_state:
        st.session_state.user_dict_applied_hash = ""
    if "llm_max_retries" not in st.session_state:
        st.session_state.llm_max_retries = 2
    if "llm_backoff_base_sec" not in st.session_state:
        st.session_state.llm_backoff_base_sec = 1.0
    if "llm_backoff_max_sec" not in st.session_state:
        st.session_state.llm_backoff_max_sec = 8.0
    if "llm_replay_rounds" not in st.session_state:
        st.session_state.llm_replay_rounds = 1
    if "llm_run_stats" not in st.session_state:
        st.session_state.llm_run_stats = {}
    if "max_rows_hard_limit" not in st.session_state:
        st.session_state.max_rows_hard_limit = 3000
    if "confirm_large_run" not in st.session_state:
        st.session_state.confirm_large_run = False
    if "export_data_bytes" not in st.session_state:
        st.session_state.export_data_bytes = None
    if "export_data_name" not in st.session_state:
        st.session_state.export_data_name = ""
    if "export_data_seq" not in st.session_state:
        st.session_state.export_data_seq = 0
    if "parsed_outline" not in st.session_state:
        st.session_state.parsed_outline = None
    if "column_type_map" not in st.session_state:
        st.session_state.column_type_map = {}
    if "manual_q_type_overrides" not in st.session_state:
        st.session_state.manual_q_type_overrides = {}


def apply_user_dict():
    terms = st.session_state.get("user_dict_terms", [])
    key = "|".join(sorted(terms))
    if st.session_state.get("user_dict_applied_hash") == key:
        return
    for w in terms:
        try:
            jieba.add_word(w)
        except Exception:
            pass


def build_export_workbook_bytes(adf, deep_report_text, ppt_text, segment_reports):
    wb = Workbook(write_only=True)
    ws_main = wb.create_sheet(title="明细+标签")
    for r in dataframe_to_rows(adf, index=False, header=True):
        ws_main.append(r)

    ws_report = wb.create_sheet(title="深度研究报告")
    report_lines = str(deep_report_text).splitlines() if deep_report_text else [
        "尚未生成深度研究报告。",
        "",
        "请在工具界面中点击“生成整体深度报告”按钮后，再导出以获得完整报告内容。",
    ]
    ws_report.append(["整体深度研究报告"])
    for line in report_lines:
        ws_report.append([line])

    ws_ppt = wb.create_sheet(title="PPT大纲")
    ppt_lines = str(ppt_text).splitlines() if ppt_text else [
        "尚未生成 PPT 大纲。",
        "",
        "请在工具界面中点击“生成 PPT 大纲”按钮后，再导出以获得完整 PPT 提纲。",
    ]
    ws_ppt.append(["PPT 大纲"])
    for line in ppt_lines:
        ws_ppt.append([line])

    if isinstance(segment_reports, dict) and segment_reports:
        for idx_seg, (seg_name, seg_report) in enumerate(
            sorted(segment_reports.items(), key=lambda x: str(x[0])), start=1
        ):
            sheet_label = f"报告-{seg_name}"
            sheet_name = make_safe_sheet_name(
                sheet_label, fallback_prefix="Seg", index=idx_seg
            )
            ws_seg = wb.create_sheet(title=sheet_name)
            ws_seg.append(["分群深挖报告"])
            for line in str(seg_report).splitlines():
                ws_seg.append([line])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

init_session_state()
config_locked = st.session_state.get("config_locked", False)
apply_user_dict()


# --- 2. 侧边栏：API 配置 ---
with st.sidebar:
    st.title("⚙️ 配置中心")
    st.session_state.api_key = st.text_input("LLM API Key", value=st.session_state.api_key, type="password", disabled=config_locked)
    st.session_state.base_url = st.text_input("API Base URL", value=st.session_state.base_url, disabled=config_locked)
    model_name = st.text_input(
        "选择模型",
        value=st.session_state.model_name,
        help="可手动输入任意模型名，如 deepseek-chat、deepseek-reasoner、gpt-4o-mini 等。",
        disabled=config_locked,
    ).strip()
    st.session_state.model_name = model_name
    batch_size = st.slider(
        "每次请求批大小",
        min_value=1,
        max_value=10,
        value=5,
        help="每个 LLM 请求中包含的样本数，数值越大单次请求越重，建议 1-5。",
        disabled=config_locked,
    )
    max_workers = st.slider(
        "并发线程数",
        min_value=1,
        max_value=10,
        value=5,
        help="同时并发的请求数量，建议 3-8 之间，根据模型限流适当调整。",
        disabled=config_locked,
    )
    st.session_state.llm_max_retries = st.slider(
        "单批次最大重试次数",
        min_value=0,
        max_value=5,
        value=st.session_state.llm_max_retries,
        help="单个批次调用 LLM 失败时的最大重试次数。",
        disabled=config_locked,
    )
    st.session_state.llm_backoff_base_sec = st.slider(
        "重试退避基准秒数",
        min_value=0.5,
        max_value=5.0,
        value=float(st.session_state.llm_backoff_base_sec),
        step=0.5,
        help="第 N 次重试等待约 base*2^N 秒，并附加随机抖动。",
        disabled=config_locked,
    )
    st.session_state.llm_backoff_max_sec = st.slider(
        "重试退避上限秒数",
        min_value=1.0,
        max_value=30.0,
        value=float(st.session_state.llm_backoff_max_sec),
        step=1.0,
        help="单次重试等待时间上限。",
        disabled=config_locked,
    )
    st.session_state.llm_replay_rounds = st.slider(
        "失败批次重放轮数",
        min_value=0,
        max_value=3,
        value=st.session_state.llm_replay_rounds,
        help="首轮并发后，对失败批次进行额外重放的轮数。",
        disabled=config_locked,
    )
    st.session_state.max_rows_hard_limit = st.number_input(
        "大样本硬阈值（行）",
        min_value=500,
        max_value=50000,
        value=int(st.session_state.max_rows_hard_limit),
        step=500,
        help="超过该阈值将要求强制确认，防止误触导致大量 API 消耗。",
        disabled=config_locked,
    )
    st.caption("—— 高级分析功能 ——")
    st.session_state.enable_emotion_analysis = st.toggle(
        "开启交互式情感趋势分析",
        value=st.session_state.enable_emotion_analysis,
        help="开启后将生成情感波动曲线图，并在深度报告中包含情感归因分析。",
        disabled=config_locked,
    )
    st.session_state.enable_text_filter = st.checkbox(
        "开启低质量文本自动过滤",
        value=st.session_state.enable_text_filter,
        disabled=config_locked,
    )
    if st.session_state.enable_text_filter:
        st.session_state.min_valid_text_length = st.slider(
            "最小有效字数限制",
            min_value=1,
            max_value=20,
            value=st.session_state.min_valid_text_length,
            disabled=config_locked,
        )
    if st.button("⏹ 终止分析"):
        st.session_state.stop_analysis = True
        st.warning("已请求终止分析：当前正在运行的分析需在终端中手动停止，新的分析将不会启动。")
    if config_locked:
        st.warning("当前分析配置已锁定。如需重新配置，请点击下方“重置并重新开始”。")
        if st.button("🔁 重置并重新开始"):
            st.session_state.config_locked = False
            st.session_state.locked_config = {}
            st.session_state.analyzed_df = None
            st.session_state.deep_report = ""
            st.session_state.ppt_outline = ""
            st.session_state.segment_reports = {}
            st.session_state.keyword_group_stats_df = None
            st.session_state.keyword_struct_for_prompt = {}
            st.session_state.keyword_stats_global = {}
            st.session_state.keyword_terms_text = ""
            st.session_state.keyword_terms_confirmed = []
            st.session_state.target_cols = []
            st.session_state.core_segment_col = None
            st.session_state.stop_analysis = False
            st.session_state.llm_run_stats = {}
            st.session_state.export_data_bytes = None
            st.session_state.export_data_name = ""
            st.rerun()
    
    st.divider()
    st.subheader("🧩 业务词典配置")
    user_dict_text = st.text_area(
        "自定义专业词汇 (防止分词切碎)",
        value=st.session_state.user_dict_text,
        help="输入游戏特有术语，用逗号、空格或换行分隔。例如：霸体, 僵直",
        height=100,
        disabled=config_locked,
    )
    st.session_state.user_dict_text = user_dict_text
    if user_dict_text:
        raw_terms = re.split(r"[,\，\s\n]+", user_dict_text)
        custom_terms = []
        for t in raw_terms:
            t_clean = t.strip()
            if t_clean and t_clean not in custom_terms:
                custom_terms.append(t_clean)
        if len(custom_terms) > 200:
            custom_terms = custom_terms[:200]
        st.session_state.user_dict_terms = custom_terms
        apply_user_dict()
    else:
        st.session_state.user_dict_terms = []
    st.caption(
        "注：深度报告中的“核心诉求”统计严格基于你选中的【分析目标列】所生成的关键词与原话池，"
        "不会把背景列或画像字段中的结构化信息当作玩家诉求来源。"
    )
    st.subheader("🎯 研究目标设置")
    st.session_state.research_goal = st.text_area(
        "本次研究目标",
        value=st.session_state.research_goal,
        height=80,
        disabled=config_locked,
    )

# --- 3. 核心功能区 ---
st.title("🧠 通用问卷文本分析引擎")
st.markdown("基于 AI 的文本分析与深度洞察引擎，支持跨列背景关联分析。")

uploaded_file = st.file_uploader("上传调研数据 (Excel / CSV / SAV)", type=["xlsx", "xls", "csv", "sav"])
outline_file = st.file_uploader(
    "上传问卷大纲（可选，.docx / .txt）",
    type=["docx", "txt"],
    help="建议上传，用于统一题型识别口径并优化目标列选择体验。",
)
outline_source = st.selectbox(
    "大纲来源",
    options=OUTLINE_PLATFORM_OPTIONS,
    index=0,
    help="解析规则按来源选择，与扩展名解耦。",
)

if uploaded_file:
    try:
        name = (uploaded_file.name or "").lower()
        if name.endswith(".xlsx") or name.endswith(".xls"):
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            if len(sheet_names) > 1:
                selected_sheet = st.selectbox(
                    "请选择要分析的工作表 (Sheet)",
                    sheet_names,
                    key="text_sheet_selector",
                )
            else:
                selected_sheet = sheet_names[0]
            df = read_table_auto(xls, sheet_name=selected_sheet)
        else:
            df = read_table_auto(uploaded_file)
        df, wjx_modified = normalize_wjx_headers(df)
        if wjx_modified:
            st.info("已自动规范化问卷星表头，便于多选/矩阵题识别。")
        df = df.reset_index(drop=True)
        if "Internal_ID" not in df.columns:
            df.insert(0, "Internal_ID", df.index + 1)
        parsed_outline = None
        if outline_file is not None:
            try:
                parsed_outline = parse_uploaded_outline_file(outline_file, outline_source)
                st.info(f"已解析问卷大纲：{len(parsed_outline)} 道题（来源：{outline_source}）")
            except Exception as e:
                st.warning(f"大纲解析失败，已回退自动识别：{e}")
                parsed_outline = None
        column_type_map = build_text_tool_column_type_map(df, outline=parsed_outline)
        st.session_state.df = df
        st.session_state.parsed_outline = parsed_outline
        st.session_state.column_type_map = column_type_map
        st.session_state.manual_q_type_overrides = {}
        st.session_state.export_data_bytes = None
        st.session_state.export_data_name = ""
        st.success(f"成功读取文件: {uploaded_file.name}, 共 {len(df)} 行数据")
    except Exception as e:
        st.error(f"读取文件失败: {e}")

if st.session_state.df is not None:
    df = st.session_state.df
    cols = df.columns.tolist()
    col_type_map = st.session_state.get("column_type_map", {}) or {}
    if not col_type_map:
        col_type_map = build_text_tool_column_type_map(df, outline=st.session_state.get("parsed_outline"))
    col_type_map = apply_manual_q_type_overrides(
        df,
        col_type_map,
        st.session_state.get("manual_q_type_overrides", {}),
    )
    st.session_state.column_type_map = col_type_map
    q_data = parse_columns_for_questions(df.columns.tolist())

    with st.expander("🛠️ 题型识别手动调整（可选）", expanded=False):
        st.caption("当自动识别有误时，可按题号手动指定题型（如将误判的单选改为“开放文本”）。")
        manual_q_type_options = ["开放文本", "单选", "多选", "矩阵", "评分", "NPS"]
        def _norm_qnum_for_cmp(v: str | int | None) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s:
                return ""
            try:
                return str(int(s))
            except ValueError:
                return s
        question_labels: list[str] = []
        for q_num in sorted(q_data.keys()):
            q_cols = [
                c
                for c in cols
                if _norm_qnum_for_cmp(extract_qnum(str(c))) == _norm_qnum_for_cmp(q_num)
            ]
            if not q_cols:
                continue
            q_type_candidates = [col_type_map.get(str(c), "未识别") for c in q_cols]
            q_type_candidates = [t for t in q_type_candidates if t not in ("元数据", "未识别")]
            q_type = q_type_candidates[0] if q_type_candidates else col_type_map.get(str(q_cols[0]), "未识别")
            stem = str(q_data[q_num].get("stem") or q_cols[0]).strip()
            question_labels.append(f"Q{int(q_num):03d} [{q_type}] {stem}")
        search_kw = st.text_input(
            "搜索题目（题号/关键词）",
            value="",
            placeholder="例如：16 或 战场信息清晰度",
            disabled=config_locked,
        ).strip().lower()
        filtered_question_labels = question_labels
        if search_kw:
            filtered_question_labels = [q for q in question_labels if search_kw in q.lower()]
        selected_q_labels = st.multiselect(
            "选择要调整的题目（可多选）",
            options=filtered_question_labels,
            default=[],
            disabled=config_locked,
        )
        selected_q_nums: list[str] = []
        for q_label in selected_q_labels:
            m_qnum = re.search(r"^Q(\d+)\s", q_label)
            q_num = m_qnum.group(1) if m_qnum else ""
            try:
                q_num = str(int(q_num))
            except ValueError:
                continue
            selected_q_nums.append(q_num)
        selected_q_nums = list(dict.fromkeys(selected_q_nums))

        forced_type = st.selectbox(
            "将所选题目统一标记为",
            options=manual_q_type_options,
            index=0,
            disabled=config_locked or (len(selected_q_nums) == 0),
        )
        c_apply, c_reset = st.columns(2)
        with c_apply:
            if st.button("批量应用题型调整", disabled=config_locked or (len(selected_q_nums) == 0)):
                overrides = dict(st.session_state.get("manual_q_type_overrides", {}))
                for q_num in selected_q_nums:
                    overrides[q_num] = forced_type
                st.session_state.manual_q_type_overrides = overrides
                st.session_state.column_type_map = apply_manual_q_type_overrides(df, col_type_map, overrides)
                st.success(f"已将 {len(selected_q_nums)} 道题标记为：{forced_type}")
                st.rerun()
        with c_reset:
            if st.button("批量恢复自动识别", disabled=config_locked or (len(selected_q_nums) == 0)):
                overrides = dict(st.session_state.get("manual_q_type_overrides", {}))
                changed = 0
                for q_num in selected_q_nums:
                    if q_num in overrides:
                        overrides.pop(q_num, None)
                        changed += 1
                if changed > 0:
                    st.session_state.manual_q_type_overrides = overrides
                    st.session_state.column_type_map = apply_manual_q_type_overrides(df, col_type_map, overrides)
                    st.success(f"已恢复 {changed} 道题自动识别")
                    st.rerun()
                else:
                    st.info("所选题目当前没有手动覆盖。")
        active_overrides = st.session_state.get("manual_q_type_overrides", {})
        if active_overrides:
            active_desc = ", ".join([f"Q{int(k):03d}→{v}" for k, v in sorted(active_overrides.items(), key=lambda x: int(x[0]))])
            st.info(f"当前手动题型调整：{active_desc}")
            if st.button("清空全部手动题型调整", disabled=config_locked):
                st.session_state.manual_q_type_overrides = {}
                st.session_state.column_type_map = build_text_tool_column_type_map(
                    df, outline=st.session_state.get("parsed_outline")
                )
                st.success("已清空全部手动题型调整。")
                st.rerun()

    target_default_options = [c for c in cols if col_type_map.get(c) == "开放文本"]
    if not target_default_options:
        target_default_options = [c for c in cols if col_type_map.get(c) not in ("元数据", "多选", "矩阵")]
    context_default_options = [c for c in cols if col_type_map.get(c) not in ("元数据", "开放文本")]
    question_map = build_question_selector_map(df, col_type_map)
    target_only_open_text = st.toggle(
        "目标列仅显示开放文本题（推荐）",
        value=True,
        help="开启后，目标列仅展示识别为开放文本的题目；关闭后可查看全部列。",
        disabled=config_locked,
    )
    question_level_mode = st.toggle(
        "按题选择（同题多列自动归并）",
        value=True,
        help="开启后，你选择的是“题目”而不是“单列”；多选题子列会自动归并为同一道题。",
        disabled=config_locked,
    )
    target_options = target_default_options if target_only_open_text else cols

    def _fmt_col(c: str) -> str:
        t = col_type_map.get(c, "未识别")
        return f"[{t}] {c}"

    def _expand_selected_keys(selected_keys: list[str], mapping: dict[str, list[str]]) -> list[str]:
        expanded: list[str] = []
        for k in selected_keys:
            expanded.extend(mapping.get(k, []))
        # 去重且保序
        return list(dict.fromkeys(expanded))

    def _key_type(k: str) -> str:
        m = re.search(r"\[(.+?)\]", k)
        return m.group(1) if m else "未识别"

    if question_level_mode:
        target_question_options = list(question_map.keys())
        if target_only_open_text:
            target_question_options = [k for k in target_question_options if _key_type(k) == "开放文本"]
        context_question_options = [k for k in question_map.keys() if _key_type(k) != "元数据"]
    target_question_keys: list[str] = []
    
    col1, col2 = st.columns(2)
    with col1:
        if question_level_mode:
            target_question_keys = st.multiselect(
                "🎯 选择分析目标题目 (文本内容)",
                options=target_question_options,
                help="按题目选择，工具会自动展开到对应列。",
                disabled=config_locked,
            )
            target_cols = _expand_selected_keys(target_question_keys, question_map)
        else:
            target_cols = st.multiselect(
                "🎯 选择分析目标列 (文本内容)",
                options=target_options,
                default=[c for c in target_default_options if c in target_options][:3],
                format_func=_fmt_col,
                help="建议优先选择开放文本题；大纲已接入时会按题型标签展示。",
                disabled=config_locked,
            )
    with col2:
        if question_level_mode:
            context_question_keys = st.multiselect(
                "💡 选择背景参考题目 (关联信息)",
                options=[k for k in context_question_options if k not in target_question_keys],
                help="按题目选择，工具会自动展开到对应列；可与目标题分离选择。",
                disabled=config_locked,
            )
            context_cols = [c for c in _expand_selected_keys(context_question_keys, question_map) if c not in target_cols]
        else:
            context_cols = st.multiselect(
                "💡 选择背景参考列 (关联信息)",
                options=[c for c in context_default_options if c not in target_cols],
                format_func=_fmt_col,
                help="建议选择结构化背景列（单选/多选/评分），避免与目标文本题同构。",
                disabled=config_locked,
            )
        st.info(
            "💡 专家建议：如何选择背景参考列？\n\n"
            "追求“独立洞察”（盲测）：若想验证玩家填空是否真实反映了其潜意识，"
            "请不要勾选与填空题内容高度相关的问卷题目（如分析玩家对未来角色期望的填空题，"
            "则不要勾选“玩家期待玩到的角色类型标签”这种高度相关的题目）。\n\n"
            "追求“逻辑归因”：若想让 AI 结合玩家的选择来解释其深层动机，则可以勾选相关选项。\n\n"
            "推荐配置：仅勾选 [用户群体]、[操作水平] 等基础属性，这能产出最具代表性的定性分析报告。"
        )
        blind_mode = st.toggle(
            "开启盲测模式（自动识别剧透标签）",
            value=st.session_state.get("blind_mode", False),
            key="blind_mode_toggle",
            help="开启后，工具会自动识别并屏蔽与目标填空题高度同构的背景列内容，使 AI 尽量只基于玩家原始表述做判断。",
            disabled=config_locked,
        )
        st.session_state.blind_mode = blind_mode
        st.session_state.context_cols = context_cols
        if question_level_mode:
            st.caption(
                f"按题选择已展开：目标 {len(target_cols)} 列，背景 {len(context_cols)} 列。"
            )
    
    core_segment_col = st.selectbox(
        "🎯 选择核心分组列（如：用户群体/玩家类型）",
        options=[None] + cols,
        index=0,
        help="该列将作为全局业务分组使用，用于整体报告和分群深挖。为空则不强制分组。",
        disabled=config_locked,
    )
    st.session_state.core_segment_col = core_segment_col
    
    if target_cols:
        st.session_state.target_cols = target_cols
        st.divider()
        st.subheader("📝 数据预览与分析配置")
        preview_cols = target_cols + context_cols
        preview_df = df[preview_cols].head(10).copy()
        enable_filter_preview = st.session_state.enable_text_filter
        min_len_preview = st.session_state.min_valid_text_length

        def _row_valid_for_preview(row):
            raw_values = []
            for t in target_cols:
                if t in row:
                    v = sanitize_value(row[t])
                    if v is not None:
                        raw_values.append(str(v))
            raw_text = " ".join(raw_values)
            if enable_filter_preview:
                return is_valid_text(raw_text, min_len_preview)
            return True

        if not preview_df.empty:
            valid_flags = preview_df.apply(_row_valid_for_preview, axis=1)
            preview_df.insert(
                0,
                "最小字数过滤结果",
                valid_flags.map(lambda x: "通过" if x else "将被拦截"),
            )
            styled = preview_df.style.apply(
                lambda r: [
                    "color: #999999"
                    if r.get("最小字数过滤结果") == "将被拦截"
                    else ""
                    for _ in r
                ],
                axis=1,
            )
            st.dataframe(styled, use_container_width=True)
            st.caption(
                "说明：当开启“最小有效字数限制”时，预览中标记为“将被拦截”的行不会发送给 LLM，"
                "而是使用本地占位结果参与后续统计。"
            )
        keyword_stats = get_keyword_stats(df, target_cols)
        st.session_state.keyword_stats_global = keyword_stats
        with st.expander("📊 填空题关键词词频初探", expanded=False):
            rows = []
            for col_name, stats in keyword_stats.items():
                for word, count in stats:
                    rows.append({"题目": col_name, "关键词": word, "提及次数": count})
            if rows:
                stats_df = pd.DataFrame(rows)
                st.dataframe(stats_df, use_container_width=True)
            else:
                st.write("当前选中的题目中暂未统计出高频关键词。")
            global_counter = Counter()
            for col_name, stats in keyword_stats.items():
                for word, count in stats:
                    global_counter[word] += count
            default_terms = [w for w, _ in global_counter.most_common(20)]
            default_text = "，".join(default_terms)
            if not st.session_state.keyword_terms_text:
                st.session_state.keyword_terms_text = default_text
            st.caption(
                "说明：系统默认预填前 20 个高频业务词，"
                "你可以自由增删修改，最终参与统计和报告生成的关键词以文本框内容为准，"
                "目前没有强制数量上限。"
            )
            kw_text = st.text_area(
                "🔎 检出高频业务词（请确认或修改）",
                value=st.session_state.keyword_terms_text,
            )
            st.session_state.keyword_terms_text = kw_text
            if kw_text:
                raw_terms = re.split(r"[,\，\s\n]+", kw_text)
                confirmed = []
                for t in raw_terms:
                    t_clean = t.strip()
                    if t_clean and t_clean not in confirmed:
                        confirmed.append(t_clean)
                st.session_state.keyword_terms_confirmed = confirmed
            else:
                st.session_state.keyword_terms_confirmed = []
            st.info("已根据当前文本框中的关键词进行统计和分析，无需额外确认按钮。")
        
        total_rows = len(df)
        estimated_batches = int((total_rows + batch_size - 1) // batch_size)
        hard_limit_rows = int(st.session_state.max_rows_hard_limit)
        if total_rows > hard_limit_rows:
            st.warning(
                f"⚠️ 当前共 {total_rows} 行，预计发起约 {estimated_batches} 个批次请求。"
                f"超过硬阈值 {hard_limit_rows} 行，可能产生较高 Token 成本。"
            )
            st.session_state.confirm_large_run = st.checkbox(
                "我已确认风险，强制执行并承担 API 费用",
                value=bool(st.session_state.confirm_large_run),
                disabled=config_locked,
                key="confirm_large_run_checkbox",
            )
        else:
            st.session_state.confirm_large_run = False

        if st.button("🚀 开始 AI 深度分析", type="primary", disabled=config_locked):
            if total_rows > hard_limit_rows and not st.session_state.get("confirm_large_run", False):
                st.warning("请先勾选“我已确认风险，强制执行并承担 API 费用”。")
                st.stop()
            st.session_state.stop_analysis = False
            if not st.session_state.api_key:
                st.warning("请先在侧边栏配置 API Key")
            else:
                st.session_state.config_locked = True
                st.session_state.locked_config = {
                    "target_cols": list(target_cols),
                    "context_cols": list(context_cols),
                    "core_segment_col": core_segment_col,
                    "blind_mode": st.session_state.get("blind_mode", False),
                    "batch_size": batch_size,
                    "max_workers": max_workers,
                    "model_name": model_name,
                    "llm_max_retries": st.session_state.llm_max_retries,
                    "llm_backoff_base_sec": st.session_state.llm_backoff_base_sec,
                    "llm_backoff_max_sec": st.session_state.llm_backoff_max_sec,
                    "llm_replay_rounds": st.session_state.llm_replay_rounds,
                }

                current_api_key = st.session_state.api_key
                current_base_url = st.session_state.base_url
                current_model_name = model_name
                current_enable_filter = st.session_state.enable_text_filter
                current_min_length = st.session_state.min_valid_text_length
                current_llm_max_retries = int(st.session_state.llm_max_retries)
                current_backoff_base = float(st.session_state.llm_backoff_base_sec)
                current_backoff_max = float(st.session_state.llm_backoff_max_sec)
                current_replay_rounds = int(st.session_state.llm_replay_rounds)
                st.session_state.llm_run_stats = {}

                blind_mode_active = st.session_state.get("blind_mode", False)
                spoiler_cols = []
                if blind_mode_active and context_cols:
                    spoiler_cols = detect_spoiler_columns(
                        target_cols,
                        context_cols,
                        current_api_key,
                        current_base_url,
                        current_model_name,
                        max_retries=current_llm_max_retries,
                        backoff_base=current_backoff_base,
                        backoff_max=current_backoff_max,
                    )
                    if spoiler_cols and len(spoiler_cols) == len(context_cols):
                        st.info(
                            "盲测模式检测到所有背景参考列均为潜在剧透，本次将暂不屏蔽任何背景列以避免完全丢失上下文。"
                        )
                        spoiler_cols = []
                    st.session_state.spoiler_cols = spoiler_cols
                    if spoiler_cols:
                        st.warning(
                            "已自动屏蔽以下高相关背景列以确保盲测纯净度："
                            + "、".join(spoiler_cols)
                        )

                # 提示词：要求模型仅输出 JSON，不包含任何多余文字或 Markdown
                prompt = ChatPromptTemplate.from_template(
                    """
你是一名资深的用户研究专家。本次研究目标是：
{research_goal}
请你根据该研究目标，对每个样本进行针对性的定性拆解和结构化标签整理。

自适应引导（非常重要）:
- 如果研究目标与「单局体验」相关，请在分析时更关注节奏、系统表现、资源反馈等体验维度。
- 如果研究目标与「满意度」或「整体评价」相关，请更侧重归因分析，找出正向驱动力与负向痛点。

输入数据说明:
- 输入是一个 JSON 数组字符串 (变量 items_json)，其中的每个元素都是一个对象，包含字段:
  - id: 样本的唯一标识符（必须原样返回）
  - segment: 原始玩家分群标签（可能为空或缺失）
  - context: 背景信息拼接字符串（可能包含评分、选择题结果等）
  - text: 填空题文本内容拼接字符串（只包含玩家的开放式文字反馈）

输入数据 (JSON 数组字符串，变量 items_json):
{items_json}

输出数据格式示例 (必须与输入数组长度相同，不可遗漏，不可重复):
[
  {{
    "item_id": "对应输入样本的id",
    "player_segment": "...",
    "key_needs": "...",
    "qualitative_insight": "...",
    "emotion_label": "...",
    "emotion_score": 7
  }},
  ...
]

各字段含义:
- item_id: 必须原样照抄输入中该条样本的 id，不得改写、不得缺失。
- player_segment: 玩家分群标签。
  - 分群一致性规则（硬性约束）：
    - 如果输入 JSON 中的对象包含 segment 字段且该字段非空，输出的 player_segment 必须与该 segment 完全一致（包括空格和符号）。
    - 严禁对 segment 进行任何翻译、改写、添加后缀或修饰。
    - 只有当 segment 为空或为 None 时，才允许根据文本和背景信息简要推断玩家分群。
  - 严禁过度解读：
    - 严禁在原始标签后添加任何解释性文字，例如将“动作搜打撤”改为“动作搜打撤/高玩挑战”是不允许的，必须保持为“动作搜打撤”。
- key_needs: 围绕研究目标提炼的核心诉求关键词，例如“高机动 / 节奏紧凑 / 减少无效等待”，使用“/”分隔多个短语。
- qualitative_insight: 面向研究目标的定性洞察总结，用 2-3 句概括该玩家在本次研究场景下的典型体验维度、系统或内容偏好、动机结构等。
- emotion_label: 对整体体验或研究对象的情绪倾向，建议取值为 "正面" / "中性" / "负面"。
- emotion_score: 情感强度评分，1-10 的整数。数值越高，情绪越强烈（无论正面或负面）。
  - 当 context 中包含满意度或评分信息（例如“非常不满意/不满意/一般/满意/非常满意”、数值评分等）时，请将这些结构化评分作为 emotion_label 和 emotion_score 的重要证据来源，并尽量与玩家文本中的情绪表达保持一致。

严格输出要求（非常重要）:
1. 只输出一个 JSON 数组字符串。
2. 不要输出任何解释性文字。
3. 不要使用 Markdown 代码块标记（例如 ```json）。
4. 数组中的每个对象必须严格包含键: item_id, player_segment, key_needs, qualitative_insight, emotion_label, emotion_score。
5. 不允许遗漏条目，也不允许重复 item_id。
                    """
                )
                
                # 开始批处理循环
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                total = len(df)
                st.info(f"正在分析 {total} 行数据，这可能需要一些时间...")
                start_time = time.time()
                processed = 0

                # 预分配结果列表，保证与原 DataFrame 行对齐
                results = [None] * total

                def process_batch(start_idx, end_idx, api_key, base_url, model_name_local, core_segment_col_local, enable_filter_local, min_length_local, context_cols_local, spoiler_cols_local, target_cols_local, research_goal_local, llm_max_retries_local, backoff_base_local, backoff_max_local):
                    batch_df = df.iloc[start_idx:end_idx]
                    items = []
                    mapping = []
                    effective_context_cols = [
                        c for c in context_cols_local if c not in spoiler_cols_local
                    ]
                    for row_idx, row in batch_df.iterrows():
                        context_data = build_context_string(row, effective_context_cols)
                        target_text = build_target_string(row, target_cols_local)
                        internal_id = sanitize_value(row.get("Internal_ID", None))
                        if internal_id is None:
                            item_id = f"row_{int(row_idx)}"
                        else:
                            item_id = str(internal_id)
                        raw_values = []
                        for t in target_cols_local:
                            if t in row:
                                v = sanitize_value(row[t])
                                if v is not None:
                                    raw_values.append(str(v))
                        raw_text = " ".join(raw_values)
                        segment_val = None
                        if core_segment_col_local and core_segment_col_local in row:
                            segment_val = sanitize_value(row[core_segment_col_local])
                        if enable_filter_local and not is_valid_text(raw_text, min_length_local):
                            placeholder = {
                                "player_segment": "无效样本",
                                "key_needs": "无",
                                "qualitative_insight": "无意义文本，已跳过",
                                "emotion_label": "中性",
                                "emotion_score": 0,
                            }
                            mapping.append({"type": "local", "value": placeholder, "segment": segment_val, "item_id": item_id})
                        else:
                            items.append({"id": item_id, "segment": segment_val, "context": context_data, "text": target_text})
                            mapping.append({"type": "llm", "item_id": item_id, "segment": segment_val})

                    llm_results = []
                    llm_success = True
                    llm_retry_count = 0
                    llm_error = ""
                    llm_raw_preview = ""
                    llm_expected_items = int(len(items))
                    llm_returned_items = 0
                    llm_matched_items = 0
                    llm_missing_items = 0
                    llm_duplicate_ids = 0
                    llm_unknown_ids = 0
                    llm_parse_failures = 0
                    if items:
                        sent_ids = [str(it["id"]) for it in items]
                        sent_id_set = set(sent_ids)
                        messages = prompt.format_messages(
                            items_json=json.dumps(items, ensure_ascii=False),
                            research_goal=research_goal_local,
                        )
                        try:
                            response, used_retry = invoke_llm_with_retry(
                                messages,
                                model_name=model_name_local,
                                api_key=api_key,
                                base_url=base_url,
                                temperature=0.3,
                                timeout=45,
                                max_retries=llm_max_retries_local,
                                backoff_base=backoff_base_local,
                                backoff_max=backoff_max_local,
                            )
                            llm_retry_count = int(used_retry)
                            raw_llm_text = response.content
                            llm_raw_preview = str(raw_llm_text)[:500]
                            data = safe_json_parse(raw_llm_text)
                            if isinstance(data, dict):
                                data_list = [data]
                            else:
                                data_list = data
                            llm_returned_items = int(len(data_list)) if isinstance(data_list, list) else 0
                            parsed_map = {}
                            for obj in data_list:
                                try:
                                    parsed_obj = UniversalAnalysis.model_validate(obj)
                                    parsed_dict = parsed_obj.model_dump()
                                    pid = str(parsed_dict.get("item_id", "")).strip()
                                    if not pid:
                                        llm_parse_failures += 1
                                        continue
                                    if pid in parsed_map:
                                        llm_duplicate_ids += 1
                                    if pid not in sent_id_set:
                                        llm_unknown_ids += 1
                                    parsed_map[pid] = parsed_dict
                                except Exception:
                                    llm_parse_failures += 1

                            llm_result_map = {}
                            for sid in sent_ids:
                                if sid in parsed_map:
                                    res_dict = parsed_map[sid].copy()
                                    res_dict.pop("item_id", None)
                                    llm_result_map[sid] = res_dict
                                    llm_matched_items += 1
                                else:
                                    llm_result_map[sid] = ERROR_PLACEHOLDER.copy()
                            llm_missing_items = max(0, llm_expected_items - llm_matched_items)
                            llm_results = [llm_result_map[sid] for sid in sent_ids]
                        except Exception as e:
                            llm_success = False
                            llm_error = str(e)
                            llm_missing_items = llm_expected_items
                            llm_results = [ERROR_PLACEHOLDER.copy() for _ in range(len(items))]

                    batch_results = []
                    for m in mapping:
                        seg_orig = m.get("segment", None)
                        seg_orig_clean = sanitize_value(seg_orig) if seg_orig is not None else None
                        if m["type"] == "local":
                            res = m["value"].copy()
                        else:
                            idx_item = next((i for i, _it in enumerate(items) if _it["id"] == m["item_id"]), -1)
                            if idx_item < len(llm_results):
                                res = llm_results[idx_item].copy()
                            else:
                                res = {"错误": "返回结果条数不足"}
                        if seg_orig_clean is not None:
                            res["player_segment"] = str(seg_orig_clean)
                        batch_results.append(res)

                    meta = {
                        "start_idx": int(start_idx),
                        "end_idx": int(end_idx),
                        "batch_rows": int(end_idx - start_idx),
                        "llm_items": int(len(items)),
                        "llm_success": bool(llm_success),
                        "llm_retry_count": int(llm_retry_count),
                        "llm_error": llm_error[:300],
                        "llm_raw_preview": llm_raw_preview,
                        "llm_expected_items": llm_expected_items,
                        "llm_returned_items": int(llm_returned_items),
                        "llm_matched_items": int(llm_matched_items),
                        "llm_missing_items": int(llm_missing_items),
                        "llm_duplicate_ids": int(llm_duplicate_ids),
                        "llm_unknown_ids": int(llm_unknown_ids),
                        "llm_parse_failures": int(llm_parse_failures),
                    }
                    return start_idx, end_idx, batch_results, meta

                batch_indexes = []
                for start_idx in range(0, total, batch_size):
                    end_idx = min(start_idx + batch_size, total)
                    batch_indexes.append((start_idx, end_idx))
                failed_batches = []
                batch_meta_records = []

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = {
                        executor.submit(
                            process_batch,
                            s,
                            e,
                            current_api_key,
                            current_base_url,
                            current_model_name,
                            core_segment_col,
                            current_enable_filter,
                            current_min_length,
                            context_cols,
                            spoiler_cols,
                            target_cols,
                            st.session_state.research_goal,
                            current_llm_max_retries,
                            current_backoff_base,
                            current_backoff_max,
                        ): (s, e)
                        for (s, e) in batch_indexes
                    }
                    for future in as_completed(futures):
                        if st.session_state.stop_analysis:
                            status_text.text("用户已请求终止分析，停止接收新的批次结果。")
                            break
                        s, e = futures[future]
                        try:
                            s, e, batch_results, meta = future.result()
                        except Exception as e_future:
                            batch_results = [ERROR_PLACEHOLDER.copy() for _ in range(max(0, e - s))]
                            meta = {
                                "start_idx": int(s),
                                "end_idx": int(e),
                                "batch_rows": int(e - s),
                                "llm_items": int(e - s),
                                "llm_success": False,
                                "llm_retry_count": 0,
                                "llm_error": str(e_future)[:300],
                            }
                        for offset, res in enumerate(batch_results):
                            idx = s + offset
                            if idx < total:
                                results[idx] = res
                        batch_meta_records.append(meta)
                        if (not meta.get("llm_success", True)) and meta.get("llm_items", 0) > 0:
                            failed_batches.append((s, e))
                        processed += (e - s)
                        progress = processed / total
                        elapsed = time.time() - start_time
                        est_total = elapsed / progress if progress > 0 else 0
                        remaining = max(0, est_total - elapsed)
                        progress_bar.progress(progress)
                        status_text.text(
                            f"已完成: {processed}/{total} 行，预计剩余 {remaining/60:.1f} 分钟"
                        )

                unresolved_batches = list(failed_batches)
                replay_total_attempts = 0
                if (not st.session_state.stop_analysis) and unresolved_batches and current_replay_rounds > 0:
                    for replay_round in range(1, current_replay_rounds + 1):
                        if not unresolved_batches:
                            break
                        status_text.text(
                            f"正在重放失败批次：第 {replay_round}/{current_replay_rounds} 轮，待重放 {len(unresolved_batches)} 批"
                        )
                        replay_total_attempts += len(unresolved_batches)
                        with ThreadPoolExecutor(max_workers=max_workers) as executor:
                            replay_futures = {
                                executor.submit(
                                    process_batch,
                                    s,
                                    e,
                                    current_api_key,
                                    current_base_url,
                                    current_model_name,
                                    core_segment_col,
                                    current_enable_filter,
                                    current_min_length,
                                    context_cols,
                                    spoiler_cols,
                                    target_cols,
                                    st.session_state.research_goal,
                                    current_llm_max_retries,
                                    current_backoff_base,
                                    current_backoff_max,
                                ): (s, e)
                                for (s, e) in unresolved_batches
                            }
                            next_unresolved = []
                            for future in as_completed(replay_futures):
                                s, e = replay_futures[future]
                                try:
                                    s, e, batch_results, meta = future.result()
                                except Exception as e_future:
                                    batch_results = [ERROR_PLACEHOLDER.copy() for _ in range(max(0, e - s))]
                                    meta = {
                                        "start_idx": int(s),
                                        "end_idx": int(e),
                                        "batch_rows": int(e - s),
                                        "llm_items": int(e - s),
                                        "llm_success": False,
                                        "llm_retry_count": 0,
                                        "llm_error": str(e_future)[:300],
                                    }
                                for offset, res in enumerate(batch_results):
                                    idx = s + offset
                                    if idx < total:
                                        results[idx] = res
                                batch_meta_records.append(meta)
                                if (not meta.get("llm_success", True)) and meta.get("llm_items", 0) > 0:
                                    next_unresolved.append((s, e))
                            unresolved_batches = next_unresolved
                
                # 合并结果
                analyzed_df = df.copy()
                results_df = pd.DataFrame(results)
                results_df = results_df.rename(columns={
                    "player_segment": "玩家分群",
                    "key_needs": "核心诉求关键词",
                    "qualitative_insight": "定性洞察",
                    "emotion_label": "情绪倾向",
                    "emotion_score": "情感强度评分",
                })
                analyzed_df = pd.concat([analyzed_df, results_df], axis=1)
                st.session_state.analyzed_df = analyzed_df
                st.session_state.export_data_bytes = None
                st.session_state.export_data_name = ""
                total_llm_rows = int(sum(m.get("llm_items", 0) for m in batch_meta_records))
                total_retry_count = int(sum(m.get("llm_retry_count", 0) for m in batch_meta_records))
                failed_batch_count = int(sum(1 for m in batch_meta_records if not m.get("llm_success", True)))
                unresolved_batch_count = int(len(unresolved_batches))
                expected_items = int(sum(m.get("llm_expected_items", 0) for m in batch_meta_records))
                returned_items = int(sum(m.get("llm_returned_items", 0) for m in batch_meta_records))
                matched_items = int(sum(m.get("llm_matched_items", 0) for m in batch_meta_records))
                missing_items = int(sum(m.get("llm_missing_items", 0) for m in batch_meta_records))
                duplicate_ids = int(sum(m.get("llm_duplicate_ids", 0) for m in batch_meta_records))
                unknown_ids = int(sum(m.get("llm_unknown_ids", 0) for m in batch_meta_records))
                parse_failures = int(sum(m.get("llm_parse_failures", 0) for m in batch_meta_records))
                st.session_state.llm_run_stats = {
                    "total_rows": int(total),
                    "llm_rows": total_llm_rows,
                    "failed_batches": failed_batch_count,
                    "unresolved_batches": unresolved_batch_count,
                    "replay_attempted_batches": int(replay_total_attempts),
                    "total_retries": total_retry_count,
                    "expected_items": expected_items,
                    "returned_items": returned_items,
                    "matched_items": matched_items,
                    "missing_items": missing_items,
                    "duplicate_ids": duplicate_ids,
                    "unknown_ids": unknown_ids,
                    "parse_failures": parse_failures,
                }
                
                st.success("分析完成！")
                stats_msg = st.session_state.llm_run_stats
                if stats_msg.get("failed_batches", 0) > 0:
                    st.info(
                        f"LLM 调用统计：重试 {stats_msg['total_retries']} 次，失败批次 {stats_msg['failed_batches']}，"
                        f"重放批次数 {stats_msg['replay_attempted_batches']}，未恢复批次 {stats_msg['unresolved_batches']}。"
                    )
                else:
                    st.info(f"LLM 调用统计：重试 {stats_msg['total_retries']} 次，全部批次成功。")
                if stats_msg.get("missing_items", 0) > 0 or stats_msg.get("duplicate_ids", 0) > 0:
                    st.warning(
                        f"对齐校验告警：期望 {stats_msg.get('expected_items', 0)}，返回 {stats_msg.get('returned_items', 0)}，"
                        f"匹配 {stats_msg.get('matched_items', 0)}，缺失 {stats_msg.get('missing_items', 0)}，"
                        f"重复ID {stats_msg.get('duplicate_ids', 0)}，未知ID {stats_msg.get('unknown_ids', 0)}。"
                    )

# --- 5. 结果展示与统计 ---
if st.session_state.analyzed_df is not None:
    adf = st.session_state.analyzed_df
    run_stats = st.session_state.get("llm_run_stats", {})
    if run_stats:
        st.caption(
            f"本轮调用统计：样本 {run_stats.get('total_rows', 0)}，LLM样本 {run_stats.get('llm_rows', 0)}，"
            f"重试 {run_stats.get('total_retries', 0)}，失败批次 {run_stats.get('failed_batches', 0)}，"
            f"重放批次数 {run_stats.get('replay_attempted_batches', 0)}，未恢复批次 {run_stats.get('unresolved_batches', 0)}，"
            f"对齐匹配 {run_stats.get('matched_items', 0)}/{run_stats.get('expected_items', 0)}。"
        )
    
    st.divider()
    st.subheader("📊 分析结果预览")
    display_df = adf.copy()
    if "Internal_ID" in display_df.columns:
        display_df = display_df.rename(columns={"Internal_ID": "样本ID"})
    if not st.session_state.get("enable_emotion_analysis", True):
        display_df = display_df.drop(columns=["情绪倾向", "情感强度评分"], errors="ignore")
    st.dataframe(display_df, use_container_width=True)
    
    tab_basic, tab_viz = st.tabs(["基础统计看板", "可视化看板"])

    with tab_basic:
        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.write("**情绪倾向分布**")
            if "情绪倾向" in adf.columns:
                fig_emo = px.pie(adf, names="情绪倾向", hole=0.3)
                st.plotly_chart(fig_emo, use_container_width=True)
        with col_stat2:
            st.write("**核心诉求 Top 关键词**")
            if "核心诉求关键词" in adf.columns:
                all_selling_points = adf["核心诉求关键词"].str.split("/|，|,").explode().str.strip()
                counts = all_selling_points.value_counts().head(10).reset_index()
                counts.columns = ["卖点", "频次"]
                fig_bar = px.bar(counts, x="卖点", y="频次", color="频次")
                st.plotly_chart(fig_bar, use_container_width=True)

    with tab_viz:
        st.markdown("### 📈 玩家情感波动分析")
        if not st.session_state.get("enable_emotion_analysis", True):
            st.info("已关闭交互式情感趋势分析。如需查看情感波动曲线，请在侧边栏开启该功能。")
        elif "情感强度评分" not in adf.columns:
            st.info("当前结果中尚未包含情感强度评分列，无法绘制情感趋势。请先完成 Pass 1 分析。")
        else:
            mode = st.radio(
                "情感视图模式",
                ["按样本序号（个体分布流）", "按维度分群（横向对比流）"],
                horizontal=True,
            )
            group_col_candidates = []
            core_seg = st.session_state.get("core_segment_col", None)
            if core_seg and core_seg in adf.columns:
                group_col_candidates.append(core_seg)
            context_cols_saved = st.session_state.get("context_cols", [])
            for c in context_cols_saved:
                if c in adf.columns and c not in group_col_candidates:
                    group_col_candidates.append(c)

            if mode == "按样本序号（个体分布流）":
                plot_df = adf.reset_index(drop=True).copy()
                plot_df["样本序号"] = plot_df.index + 1
                hover_cols = {}
                if "定性洞察" in plot_df.columns:
                    hover_cols["定性洞察"] = True
                if "核心诉求关键词" in plot_df.columns:
                    hover_cols["核心诉求关键词"] = True
                x_field_options = ["样本序号（默认）"]
                time_like_cols = []
                for col_name, dtype in adf.dtypes.items():
                    dtype_str = str(dtype)
                    if dtype_str.startswith(("int", "float", "datetime")):
                        if col_name != "情感强度评分":
                            x_field_options.append(col_name)
                    elif dtype_str.startswith("object"):
                        name_lower = str(col_name).lower()
                        if ("时间" in str(col_name)) or ("time" in name_lower) or ("date" in name_lower):
                            time_like_cols.append(col_name)
                for col_name in time_like_cols:
                    try:
                        parsed = pd.to_datetime(adf[col_name], errors="coerce")
                        if parsed.notna().sum() > 0:
                            plot_df[col_name] = parsed
                            if col_name not in x_field_options:
                                x_field_options.append(col_name)
                    except Exception:
                        pass
                x_choice = st.selectbox(
                    "选择情感趋势的 X 轴字段",
                    options=x_field_options,
                    index=0,
                )
                if x_choice == "样本序号（默认）":
                    x_field = "样本序号"
                else:
                    x_field = x_choice
                max_points = 3000
                if len(plot_df) > max_points:
                    plot_df = plot_df.iloc[:max_points]
                fig_scatter = px.scatter(
                    plot_df,
                    x=x_field,
                    y="情感强度评分",
                    hover_data=hover_cols,
                )
                fig_scatter.update_traces(marker=dict(size=6, opacity=0.6))
                fig_scatter.update_layout(xaxis_title=x_field, yaxis_title="情感强度评分")
                st.plotly_chart(fig_scatter, use_container_width=True)
            else:
                if not group_col_candidates:
                    st.info("当前尚未选择核心分组或背景列，无法按维度对比分组的情感分布。")
                else:
                    group_col = st.selectbox(
                        "选择分群维度",
                        options=group_col_candidates,
                        index=0,
                    )
                    box_df = adf[[group_col, "情感强度评分"]].copy()
                    if "定性洞察" in adf.columns:
                        box_df["定性洞察"] = adf["定性洞察"]
                    if "核心诉求关键词" in adf.columns:
                        box_df["核心诉求关键词"] = adf["核心诉求关键词"]
                    box_df = box_df.dropna(subset=["情感强度评分"])
                    hover_cols = {}
                    if "定性洞察" in box_df.columns:
                        hover_cols["定性洞察"] = True
                    if "核心诉求关键词" in box_df.columns:
                        hover_cols["核心诉求关键词"] = True
                    points_mode = "all" if len(box_df) <= 2000 else "outliers"
                    fig_box = px.box(
                        box_df,
                        x=group_col,
                        y="情感强度评分",
                        points=points_mode,
                        hover_data=hover_cols,
                    )
                    fig_box.update_layout(
                        xaxis_title=group_col,
                        yaxis_title="情感强度评分",
                    )
                    st.plotly_chart(fig_box, use_container_width=True)

    core_segment_col_for_kw = st.session_state.get("core_segment_col", None)
    confirmed_keywords = st.session_state.get("keyword_terms_confirmed", [])
    target_cols_for_kw = st.session_state.get("target_cols", [])
    if (
        core_segment_col_for_kw
        and core_segment_col_for_kw in st.session_state.df.columns
        and confirmed_keywords
        and target_cols_for_kw
    ):
        kw_stats_df, kw_struct = compute_keyword_group_stats(
            st.session_state.df,
            target_cols_for_kw,
            core_segment_col_for_kw,
            confirmed_keywords,
        )
        st.session_state.keyword_group_stats_df = kw_stats_df
        st.session_state.keyword_struct_for_prompt = kw_struct
        if kw_stats_df is not None and not kw_stats_df.empty:
            st.subheader("🔍 核心业务关键词在不同分组下的关注度对比")
            agg = (
                kw_stats_df.groupby(["核心分组", "关键词"])
                .agg({"提及人数": "sum", "组样本数": "max"})
                .reset_index()
            )
            agg["提及率"] = agg["提及人数"] / agg["组样本数"].replace(0, pd.NA)
            agg["提及率百分比"] = agg["提及率"] * 100
            fig_kw = px.bar(
                agg,
                x="提及率百分比",
                y="关键词",
                color="核心分组",
                orientation="h",
                barmode="group",
                hover_data=["提及人数", "组样本数"],
            )
            fig_kw.update_layout(xaxis_title="提及率（%）")
            st.plotly_chart(fig_kw, use_container_width=True)

    # --- 5.1 深度报告生成（Pass 2） ---
    st.divider()
    st.subheader("🧠 深度研究报告（Pass 2）")
    st.markdown("在完成逐条标签化后，可对整体样本进行多级聚合与对比分析。")

    if st.button("生成整体深度报告"):
        try:
            # 仅保留结构化分析相关列，并强制保留核心分组列（如有）
            core_segment_col = st.session_state.get("core_segment_col", None)
            cols_for_summary = [
                "Internal_ID",
                "玩家分群",
                "核心诉求关键词",
                "定性洞察",
                "情绪倾向",
                "情感强度评分",
            ]
            available_cols = [c for c in cols_for_summary if c in adf.columns]
            if core_segment_col and core_segment_col in adf.columns and core_segment_col not in available_cols:
                available_cols.insert(0, core_segment_col)
            summary_df = adf[available_cols].copy()

            total_n = len(summary_df)
            # 简单采样控制上下文大小
            MAX_RECORDS = 200
            if total_n > MAX_RECORDS:
                sample_df = summary_df.sample(MAX_RECORDS, random_state=42)
                sample_note = f"当前共有 {total_n} 条样本，本次向模型发送了其中 {MAX_RECORDS} 条的代表性样本。"
            else:
                sample_df = summary_df
                sample_note = f"当前共有 {total_n} 条样本，已全部用于生成报告。"

            # 计算核心分组列的真实分布（优先使用用户选择的核心分组列）
            core_col_for_stats = None
            if core_segment_col and core_segment_col in summary_df.columns:
                core_col_for_stats = core_segment_col
            elif "玩家分群" in summary_df.columns:
                core_col_for_stats = "玩家分群"

            if core_col_for_stats:
                segment_counts = summary_df[core_col_for_stats].value_counts(dropna=False).reset_index()
                if not segment_counts.empty:
                    segment_counts.columns = ["分组标签", "样本数"]
                    segment_counts["占比"] = (segment_counts["样本数"] / total_n * 100).round(1)
                    segment_stats = segment_counts.to_dict(orient="records")
                else:
                    segment_stats = []
            else:
                segment_stats = []

            sample_for_prompt = sample_df.rename(columns={"Internal_ID": "sample_id"}) if "Internal_ID" in sample_df.columns else sample_df.copy()
            sample_records = sample_for_prompt.to_dict(orient="records")

            target_cols_for_kw = st.session_state.get("target_cols", [])
            keyword_struct = st.session_state.get("keyword_struct_for_prompt", {})
            core_segment_col_for_kw = st.session_state.get("core_segment_col", None)
            confirmed_keywords = st.session_state.get("keyword_terms_confirmed", [])
            if (
                target_cols_for_kw
                and core_segment_col_for_kw
                and core_segment_col_for_kw in st.session_state.df.columns
                and confirmed_keywords
                and not keyword_struct
            ):
                kw_stats_df, kw_struct = compute_keyword_group_stats(
                    st.session_state.df,
                    target_cols_for_kw,
                    core_segment_col_for_kw,
                    confirmed_keywords,
                )
                st.session_state.keyword_group_stats_df = kw_stats_df
                st.session_state.keyword_struct_for_prompt = kw_struct
                keyword_struct = kw_struct
            kw_stats_df_for_prompt = st.session_state.get("keyword_group_stats_df", None)

            keyword_group_stats_payload = {}
            significant_deviation_items = []
            deviation_evidence_pool = {}
            try:
                if kw_stats_df_for_prompt is not None and not kw_stats_df_for_prompt.empty:
                    df_stats = kw_stats_df_for_prompt.copy()
                    df_stats["__key__"] = df_stats["题目"].astype(str) + "||" + df_stats["关键词"].astype(str)
                    overall_map = {}
                    for key, sub in df_stats.groupby("__key__"):
                        total_mention = float(sub["提及人数"].sum())
                        total_n_keyword = float(sub["组样本数"].sum())
                        overall_rate = total_mention / total_n_keyword if total_n_keyword > 0 else 0.0
                        overall_map[key] = {
                            "overall_mention_users": int(total_mention),
                            "overall_mention_rate": overall_rate,
                        }
                    for _, row in df_stats.iterrows():
                        q = str(row["题目"])
                        kw = str(row["关键词"])
                        seg = str(row["核心分组"])
                        mention_users = int(row["提及人数"])
                        group_n = int(row["组样本数"])
                        mention_rate = float(row["提及率"])
                        key = f"{q}||{kw}"
                        overall_info = overall_map.get(key, {"overall_mention_users": 0, "overall_mention_rate": 0.0})
                        overall_rate = overall_info["overall_mention_rate"]
                        deviation = mention_rate - overall_rate
                        if abs(deviation) < 0.15:
                            continue
                        if q not in keyword_group_stats_payload:
                            keyword_group_stats_payload[q] = {}
                        if kw not in keyword_group_stats_payload[q]:
                            keyword_group_stats_payload[q][kw] = {
                                "overall_mention_users": overall_info["overall_mention_users"],
                                "overall_mention_rate": overall_rate,
                                "segments": [],
                            }
                        keyword_group_stats_payload[q][kw]["segments"].append(
                            {
                                "segment": seg,
                                "mention_users": mention_users,
                                "mention_rate": mention_rate,
                                "group_size": group_n,
                                "deviation_vs_overall": deviation,
                            }
                        )
                    for q, kw_dict in keyword_group_stats_payload.items():
                        for kw, info in kw_dict.items():
                            seg_list = info.get("segments", [])
                            if seg_list:
                                seg_list_sorted = sorted(
                                    seg_list,
                                    key=lambda s: abs(s.get("deviation_vs_overall", 0.0)),
                                    reverse=True,
                                )
                                info["segments"] = seg_list_sorted[:5]
                if (
                    target_cols_for_kw
                    and core_segment_col_for_kw
                    and core_segment_col_for_kw in st.session_state.df.columns
                    and confirmed_keywords
                ):
                    significant_deviation_items = compute_keyword_deviations(
                        st.session_state.df,
                        target_cols_for_kw,
                        core_segment_col_for_kw,
                        confirmed_keywords,
                        threshold=0.15,
                        topk=15,
                    )
                    deviation_evidence_pool = prepare_evidence_pool(
                        st.session_state.df,
                        target_cols_for_kw,
                        core_segment_col_for_kw,
                        significant_deviation_items,
                        max_quotes_per_item=5,
                    )
            except Exception:
                keyword_group_stats_payload = {}
                significant_deviation_items = []
                deviation_evidence_pool = {}

            raw_quotes_by_segment = {}
            df_raw = st.session_state.df
            max_quotes_per_segment = 20
            if target_cols_for_kw and core_col_for_stats and core_col_for_stats in df_raw.columns:
                for seg_value, g in df_raw.groupby(core_col_for_stats, dropna=False):
                    quotes = []
                    for _, row_raw in g.iterrows():
                        raw_parts = []
                        for tcol in target_cols_for_kw:
                            if tcol in row_raw:
                                v = sanitize_value(row_raw[tcol])
                                if v is not None:
                                    raw_parts.append(str(v))
                        combined = " ".join(raw_parts).strip()
                        if not combined:
                            continue
                        if st.session_state.enable_text_filter and not is_valid_text(
                            combined, st.session_state.min_valid_text_length
                        ):
                            continue
                        sid = row_raw.get("Internal_ID", None)
                        if sid is not None:
                            combined = f"[#{sid}] " + combined
                        quotes.append(combined)
                        if len(quotes) >= max_quotes_per_segment:
                            break
                    if quotes:
                        raw_quotes_by_segment[str(seg_value)] = quotes
            else:
                quotes = []
                for _, row_raw in df_raw.iterrows():
                    raw_parts = []
                    for tcol in target_cols_for_kw:
                        if tcol in row_raw:
                            v = sanitize_value(row_raw[tcol])
                            if v is not None:
                                raw_parts.append(str(v))
                    combined = " ".join(raw_parts).strip()
                    if not combined:
                        continue
                    if st.session_state.enable_text_filter and not is_valid_text(
                        combined, st.session_state.min_valid_text_length
                    ):
                        continue
                    sid = row_raw.get("Internal_ID", None)
                    if sid is not None:
                        combined = f"[#{sid}] " + combined
                    quotes.append(combined)
                    if len(quotes) >= 20:
                        break
                if quotes:
                    raw_quotes_by_segment["ALL"] = quotes

            collision_background = {}
            try:
                context_cols_saved = st.session_state.get("context_cols", [])
                if df_raw is not None and context_cols_saved:
                    for col in context_cols_saved:
                        if col not in df_raw.columns:
                            continue
                        ser_all = df_raw[col].astype(str).fillna("").str.strip()
                        overall_total = int(len(ser_all))
                        if overall_total == 0:
                            continue
                        vc_all = ser_all.value_counts().head(20)
                        col_info = {
                            "overall": [],
                            "by_segment": {},
                        }
                        for value, cnt in vc_all.items():
                            rate = float(cnt) / overall_total if overall_total > 0 else 0.0
                            col_info["overall"].append(
                                {"value": str(value), "count": int(cnt), "rate": rate}
                            )
                        if core_col_for_stats and core_col_for_stats in df_raw.columns:
                            for seg_value, g in df_raw.groupby(core_col_for_stats, dropna=False):
                                g_ser = g[col].astype(str).fillna("").str.strip()
                                seg_total = int(len(g_ser))
                                if seg_total == 0:
                                    continue
                                vc_seg = g_ser.value_counts().head(20)
                                seg_list = []
                                for value, cnt in vc_seg.items():
                                    rate = float(cnt) / seg_total if seg_total > 0 else 0.0
                                    seg_list.append(
                                        {"value": str(value), "count": int(cnt), "rate": rate}
                                    )
                                col_info["by_segment"][str(seg_value)] = {
                                    "total": seg_total,
                                    "values": seg_list,
                                }
                        collision_background[col] = col_info
            except Exception:
                collision_background = {}
            global_payload = {
                "sample_note": sample_note,
                "total_samples": total_n,
                "segment_stats": segment_stats,
                "samples": sample_records,
                "keyword_context": keyword_struct,
                "keyword_group_stats": keyword_group_stats_payload,
                "raw_quotes_by_segment": raw_quotes_by_segment,
                "collision_background": collision_background,
                "significant_keyword_deviations": significant_deviation_items,
                "deviation_evidence_pool": deviation_evidence_pool,
            }

            llm_for_report = ChatOpenAI(
                model=model_name,
                openai_api_key=st.session_state.api_key,
                openai_api_base=st.session_state.base_url,
                temperature=0.3,
            )

            enable_emotion = st.session_state.get("enable_emotion_analysis", True)
            if enable_emotion:
                report_template_text = """
你是一名专业的游戏用户研究员。你具备极强的文本洞察力，能够准确归因并提炼玩家真实诉求。
本次分析的核心研究目标是：{research_goal}。请务必以此目标为准绳，过滤无关杂音，聚焦核心信息，输出一份结构化的深度研究报告。

【数据说明】:
{sample_note}
总样本数: {total_samples}

【玩家分群统计】(segment_stats，JSON 格式):
{segment_stats_json}

【填空题高频关键词统计】(keyword_context，JSON 格式):
{keyword_context_json}

【按分组的关键词对比统计】(keyword_group_stats，JSON 格式):
{keyword_group_stats_json}

【样本结构化明细】(samples，JSON 格式，已做抽样或截断):
{samples_json}

【按分组采样的原始玩家原话】(raw_quotes_by_segment，JSON 格式):
{raw_quotes_json}

【按分组的背景标签统计】(collision_background，JSON 格式):
{collision_background_json}

【样本追溯规则】(最高优先级):
- 在本次输入的 samples 和 raw_quotes_by_segment 中，每条玩家反馈都带有唯一的 sample_id。
- 在 raw_quotes_by_segment 中，sample_id 以「[#数字]」的形式出现在每条原话的开头，例如：[#12] 代表样本 ID 为 12。
- 你在撰写报告时，严禁凭空捏造结论。每当你提出一个关键洞察或发现一个“对撞冲突”时，必须在结论后以括号形式明确引用对应的 sample_id，例如：「……（证据：见样本 #12、#47）」。
- 严禁自行编造不存在的 sample_id，只能引用输入数据中实际出现的 ID。
写作要求（请按结构分段输出）:
1. 总览结论:
   - 用 3-5 条要点，概括本次研究的整体发现。
   - 数据源分离原则：本章节中的“核心诉求相关结论”必须以 keyword_context（关键词统计）和 raw_quotes_by_segment（原始玩家原话）为首要依据。
2. 按玩家分群的横向对比:
   - 对每个玩家分群，概括其核心诉求、典型体验维度和情绪特征。
   - 在概括“核心诉求”时，必须显式引用 keyword_context 中提供的频次与占比数据，例如：
     「根据关键词统计，某关键词在某分组中被提及 XX 次，占该组样本的约 YY%，是该分组最核心的期待之一」。
   - 结合 keyword_group_stats，重点观察“群体提及率”与“大盘平均水平”的偏离情况：如果某个词在大盘提及率约为 p%，而在某分组中提及率为 (p + Δ)%，请显式指出「该分组对某关键词的关注度显著高于大盘（+Δ%）」，并给出业务含义。
   - 严禁数据回流：严禁将 samples 中 qualitative_insight 字段里的总结性描述当作“核心诉求”进行独立统计，核心诉求的量化依据必须来自 keyword_context 与 keyword_group_stats 提供的频次和偏离度。
   - 针对 keyword_group_stats 与 significant_keyword_deviations 中偏离度绝对值超过 15% 的高偏离关键词，结合 deviation_evidence_pool 中提供的原话证据，补充一小节「分群差异化诉求深度归因」，解释为什么该分群对该词的关注度异常偏高或偏低，并在每条归因结论后引用 1-2 个 sample_id 作为证据。
3. 核心体验维度与典型主题聚类:
   - 使用 qualitative_insight 与情绪表达，对玩家反馈进行主题聚类，总结 3 个左右最关键的体验维度或话题簇（例如：节奏紧凑 / 反馈清晰 / 负担过重 等）。
   - 说明每个维度下的典型正向与负向反馈，并指出与不同玩家分群之间的差异。
4. 情感趋势解读:
   - 基于 emotion_score 的整体分布和按分组拆分的差异，识别本次体验中的“波峰”（高分集中、玩家明显兴奋或满足的环节）与“波谷”（低分集中、玩家明显挫败或劝退的环节）。
   - 用少量清晰的段落，分别描述哪些体验片段对应高分峰值、哪些片段对应低分谷值，并尽量引用具体的定性线索或关键词来佐证。
5. 情绪与体验风险:
   - 在情感趋势的基础上，进一步概括高情感强度（高分和低分）背后的正向与负向因素，指出潜在的流失风险点或口碑爆点。
6. 可落地建议:
   - 列出 3-5 条可以直接给到策划/产品/运营同学的行动建议。
7. 【数据真实性校验（对撞分析）】:
   - 基于 collision_background（定量背景标签分布）与 qualitative_insight / raw_quotes_by_segment（定性文本），检查“玩家勾选的选项”和“自由文本中的表达”是否一致。
   - 重点识别“言行不一”的典型情况，例如：在定量题中勾选了“高机动/硬核”等标签，但在开放式原话中频繁抱怨“挫败感过强”“动作僵硬”等与标签不一致的体验。
   - 对于你识别出的每一个“需求错位点”或“认知冲突点”，必须至少列出 1-3 个具体的 sample_id 作为证据，引用格式示例：「……（证据：见样本 #89、#102）」。
   - 所有 sample_id 必须来自输入数据中实际出现的 ID，严禁编造。如果在某个分群或某类标签下未发现足够清晰的冲突样本，请在该部分明确写出“暂无典型冲突样本”。

关于 keyword_context（数据驱动的关键词统计）:
- keyword_context 的结构示例:
  {{
    "题目A": {{
      "关键词1": {{
        "total_mention_users": X,
        "by_segment": {{"分组1": {{"mention_users": a, "mention_rate": b}}, ...}}
      }},
      "关键词2": {{ ... }}
    }},
    "题目B": {{ ... }}
  }}

输出格式:
- 使用中文分段撰写，不要输出任何 JSON。
- 在每一条“核心诉求相关”的关键洞察下方，请从 raw_quotes_by_segment 提供的原话池中选取 3-5 条最具代表性的玩家原话，以 Markdown 引用块（> 开头）或列表的形式展示，确保“每一项结论都有原话作证”，且原话内容来自玩家开放式反馈，而非结构化画像字段。
- 在分析时，如果你的语义结论与 keyword_context 中的频次和占比一致，请在文字中明确引用相应数字进行佐证；
- 如果语义结论与 keyword_context 中的数据存在明显冲突，请在报告中指出可能原因（如样本偏差、题目歧义等），避免给出与统计显著不一致的定性判断。
- 【防幻觉与严谨性准则】:
  - 规则 1（证据缺失时保持沉默）: 如果在 raw_quotes_by_segment 或 deviation_evidence_pool 中找不到能够直接支持某个统计偏差的具体语境，请不要依靠“可能、推测、常理经验”进行归因，而是明确标注为「统计显著，但定性证据不足，暂无法直接归因」。
  - 规则 2（禁止跨数据源臆测）: 严禁用本次输入数据之外的常识或经验（例如“动作玩家通常喜欢挑战”）来解释偏差，所有归因结论都必须能在当前文本池中找到对应的语义线索。
  - 规则 3（ID 闭环校验）: 在引用 sample_id 作为证据时，必须确保对应原话中确实包含该分析所指的关键词或场景，严禁随意粘贴 ID 充数。
  - 规则 4（对撞过滤器）: 如果你的归因结论与 collision_background 中的定量背景存在明显矛盾（例如归因为“追求高难度”，但该分群在定量题中主要选择“轻度休闲”），需要在报告中点出这一矛盾点，而不是强行统一结论。
                """
            else:
                report_template_text = """
你是一名专业的游戏用户研究员。你具备极强的文本洞察力，能够准确归因并提炼玩家真实诉求。
本次分析的核心研究目标是：{research_goal}。本次分析不包含情感维度，请忽略样本中的情绪标签和情感强度，不要在报告中生成任何关于情感、满意度分数或情绪波动的内容，而是聚焦在文本中显性的诉求、体验维度和结构化差异。

【数据说明】:
{sample_note}
总样本数: {total_samples}

【玩家分群统计】(segment_stats，JSON 格式):
{segment_stats_json}

【填空题高频关键词统计】(keyword_context，JSON 格式):
{keyword_context_json}

【按分组的关键词对比统计】(keyword_group_stats，JSON 格式):
{keyword_group_stats_json}

【样本结构化明细】(samples，JSON 格式，已做抽样或截断):
{samples_json}

【按分组采样的原始玩家原话】(raw_quotes_by_segment，JSON 格式):
{raw_quotes_json}

【按分组的背景标签统计】(collision_background，JSON 格式):
{collision_background_json}

【样本追溯规则】(最高优先级):
- 在本次输入的 samples 和 raw_quotes_by_segment 中，每条玩家反馈都带有唯一的 sample_id。
- 在 raw_quotes_by_segment 中，sample_id 以「[#数字]」的形式出现在每条原话的开头，例如：[#12] 代表样本 ID 为 12。
- 你在撰写报告时，严禁凭空捏造结论。每当你提出一个关键洞察或发现一个“对撞冲突”时，必须在结论后以括号形式明确引用对应的 sample_id，例如：「……（证据：见样本 #12、#47）」。
- 严禁自行编造不存在的 sample_id，只能引用输入数据中实际出现的 ID。

写作要求（请按结构分段输出）:
1. 总览结论:
   - 用 3-5 条要点，概括本次研究的整体发现。
   - 数据源分离原则：本章节中的“核心诉求相关结论”必须以 keyword_context（关键词统计）和 raw_quotes_by_segment（原始玩家原话）为首要依据。
2. 按玩家分群的横向对比:
   - 对每个玩家分群，概括其核心诉求和典型体验维度（但不做情绪趋势或满意度高低的评价）。
   - 在概括“核心诉求”时，必须显式引用 keyword_context 中提供的频次与占比数据，例如：
     「根据关键词统计，某关键词在某分组中被提及 XX 次，占该组样本的约 YY%，是该分组最核心的期待之一」。
   - 结合 keyword_group_stats，重点观察“群体提及率”与“大盘平均水平”的偏离情况：如果某个词在大盘提及率约为 p%，而在某分组中提及率为 (p + Δ)%，请显式指出「该分组对某关键词的关注度显著高于大盘（+Δ%）」，并给出业务含义。
   - 严禁数据回流：严禁将 samples 中 qualitative_insight 字段里的总结性描述当作“核心诉求”进行独立统计，核心诉求的量化依据必须来自 keyword_context 与 keyword_group_stats 提供的频次和偏离度。
   - 针对 keyword_group_stats 与 significant_keyword_deviations 中偏离度绝对值超过 15% 的高偏离关键词，结合 deviation_evidence_pool 中提供的原话证据，补充一小节「分群差异化诉求深度归因」，解释为什么该分群对该词的关注度异常偏高或偏低，并在每条归因结论后引用 1-2 个 sample_id 作为证据。
3. 核心体验维度与典型主题聚类:
   - 使用 qualitative_insight，对玩家反馈进行主题聚类，总结 3 个左右最关键的体验维度或话题簇（例如：节奏紧凑 / 反馈清晰 / 负担过重 等）。
   - 说明每个维度下的典型正向与负向反馈，并指出与不同玩家分群之间的差异。
4. 可落地建议:
   - 列出 3-5 条可以直接给到策划/产品/运营同学的行动建议。
5. 【数据真实性校验（对撞分析）】:
   - 基于 collision_background（定量背景标签分布）与 qualitative_insight / raw_quotes_by_segment（定性文本），检查“玩家勾选的选项”和“自由文本中的表达”是否一致。
   - 重点识别“言行不一”的典型情况，例如：在定量题中勾选了某标签，但在开放式原话中频繁抱怨与该标签相反的体验。
   - 对于你识别出的每一个“需求错位点”或“认知冲突点”，必须至少列出 1-3 个具体的 sample_id 作为证据，引用格式示例：「……（证据：见样本 #89、#102）」。
   - 所有 sample_id 必须来自输入数据中实际出现的 ID，严禁编造。如果在某个分群或某类标签下未发现足够清晰的冲突样本，请在该部分明确写出“暂无典型冲突样本”。

关于 keyword_context（数据驱动的关键词统计）:
- keyword_context 的结构示例:
  {{
    "题目A": {{
      "关键词1": {{
        "total_mention_users": X,
        "by_segment": {{"分组1": {{"mention_users": a, "mention_rate": b}}, ...}}
      }},
      "关键词2": {{ ... }}
    }},
    "题目B": {{ ... }}
  }}

输出格式:
- 使用中文分段撰写，不要输出任何 JSON。
- 在每一条“核心诉求相关”的关键洞察下方，请从 raw_quotes_by_segment 提供的原话池中选取 3-5 条最具代表性的玩家原话，以 Markdown 引用块（> 开头）或列表的形式展示，确保“每一项结论都有原话作证”，且原话内容来自玩家开放式反馈，而非结构化画像字段。
- 在分析时，如果你的语义结论与 keyword_context 中的频次和占比一致，请在文字中明确引用相应数字进行佐证；
- 如果语义结论与 keyword_context 中的数据存在明显冲突，请在报告中指出可能原因（如样本偏差、题目歧义等），避免给出与统计显著不一致的定性判断。
- 【防幻觉与严谨性准则】:
-  - 规则 1（证据缺失时保持沉默）: 如果在 raw_quotes_by_segment 或 deviation_evidence_pool 中找不到能够直接支持某个统计偏差的具体语境，请不要依靠“可能、推测、常理经验”进行归因，而是明确标注为「统计显著，但定性证据不足，暂无法直接归因」。
-  - 规则 2（禁止跨数据源臆测）: 严禁用本次输入数据之外的常识或经验（例如“动作玩家通常喜欢挑战”）来解释偏差，所有归因结论都必须能在当前文本池中找到对应的语义线索。
-  - 规则 3（ID 闭环校验）: 在引用 sample_id 作为证据时，必须确保对应原话中确实包含该分析所指的关键词或场景，严禁随意粘贴 ID 充数。
-  - 规则 4（对撞过滤器）: 如果你的归因结论与 collision_background 中的定量背景存在明显矛盾（例如归因为“追求高难度”，但该分群在定量题中主要选择“轻度休闲”），需要在报告中点出这一矛盾点，而不是强行统一结论。
                """

            report_template = ChatPromptTemplate.from_template(report_template_text)

            messages = report_template.format_messages(
                sample_note=global_payload["sample_note"],
                total_samples=global_payload["total_samples"],
                segment_stats_json=json.dumps(global_payload["segment_stats"], ensure_ascii=False),
                keyword_context_json=json.dumps(global_payload["keyword_context"], ensure_ascii=False),
                samples_json=json.dumps(global_payload["samples"], ensure_ascii=False),
                raw_quotes_json=json.dumps(global_payload["raw_quotes_by_segment"], ensure_ascii=False),
                keyword_group_stats_json=json.dumps(global_payload["keyword_group_stats"], ensure_ascii=False),
                collision_background_json=json.dumps(global_payload["collision_background"], ensure_ascii=False),
                significant_deviations_json=json.dumps(global_payload["significant_keyword_deviations"], ensure_ascii=False),
                deviation_evidence_pool_json=json.dumps(global_payload["deviation_evidence_pool"], ensure_ascii=False),
                research_goal=st.session_state.research_goal,
            )

            report_response = llm_for_report.invoke(messages)
            deep_report_text = report_response.content
            st.session_state.deep_report = deep_report_text
            st.success("深度研究报告已生成。")
        except Exception as e:
            st.error(f"生成深度报告时出错: {e}")

    if "deep_report" in st.session_state and st.session_state.deep_report:
        st.markdown("#### 整体深度研究报告预览")
        st.text_area("整体深度研究报告", st.session_state.deep_report, height=280)

    core_segment_col = st.session_state.get("core_segment_col", None)
    if core_segment_col and core_segment_col in adf.columns:
        st.markdown("#### 按核心分组列进行深挖报告")
        segments = sorted(adf[core_segment_col].dropna().unique().tolist())
        if segments:
            seg_col1, seg_col2 = st.columns([2, 1])
            with seg_col1:
                selected_segments = st.multiselect(
                    f"选择一个或多个 [{core_segment_col}] 分组进行深挖",
                    options=segments,
                    key="segments_for_deep_dive",
                )
            with seg_col2:
                trigger_segment_report = st.button("批量生成深挖报告")

            if trigger_segment_report:
                if not selected_segments:
                    st.warning("请至少选择一个分组进行深挖。")
                else:
                    total_seg = len(selected_segments)
                    progress_seg = st.progress(0.0)
                    status_seg = st.empty()
                    reports = st.session_state.get("segment_reports", {})
                    for idx_seg, selected_segment in enumerate(selected_segments, start=1):
                        try:
                            group_series = adf[core_segment_col].astype(str).str.strip()
                            selected_norm = str(selected_segment).strip()
                            seg_mask = group_series == selected_norm
                            seg_df = adf[seg_mask].copy()
                            seg_n = len(seg_df)
                            if seg_n == 0:
                                reports[selected_norm] = "该分群当前没有样本。"
                            else:
                                cols_for_summary = [
                                    "玩家分群",
                                    "核心诉求关键词",
                                    "定性洞察",
                                    "情绪倾向",
                                    "情感强度评分",
                                ]
                                seg_available_cols = [c for c in cols_for_summary if c in seg_df.columns]
                                seg_summary_df = seg_df[seg_available_cols].copy()
                                MAX_SEG_RECORDS = 150
                                if seg_n > MAX_SEG_RECORDS:
                                    seg_sample_df = seg_summary_df.sample(MAX_SEG_RECORDS, random_state=42)
                                    seg_sample_note = f"当前分组 [{core_segment_col} = {selected_norm}] 有 {seg_n} 条样本，本次向模型发送了其中 {MAX_SEG_RECORDS} 条的代表性样本。"
                                else:
                                    seg_sample_df = seg_summary_df
                                    seg_sample_note = f"当前分组 [{core_segment_col} = {selected_norm}] 有 {seg_n} 条样本，已全部用于生成该分组深挖报告。"
                                seg_sample_records = seg_sample_df.to_dict(orient="records")
                                target_cols_for_kw = st.session_state.get("target_cols", [])
                                segment_keyword_stats = {}
                                if target_cols_for_kw:
                                    segment_keyword_stats = get_keyword_stats(seg_df, target_cols_for_kw)
                                segment_raw_quotes = []
                                max_quotes_per_segment = 20
                                if target_cols_for_kw:
                                    for _, row_raw in seg_df.iterrows():
                                        raw_parts = []
                                        for tcol in target_cols_for_kw:
                                            if tcol in row_raw:
                                                v = sanitize_value(row_raw[tcol])
                                                if v is not None:
                                                    raw_parts.append(str(v))
                                        combined = " ".join(raw_parts).strip()
                                        if not combined:
                                            continue
                                        if st.session_state.enable_text_filter and not is_valid_text(
                                            combined, st.session_state.min_valid_text_length
                                        ):
                                            continue
                                        segment_raw_quotes.append(combined)
                                        if len(segment_raw_quotes) >= max_quotes_per_segment:
                                            break
                                llm_for_segment = ChatOpenAI(
                                    model=model_name,
                                    openai_api_key=st.session_state.api_key,
                                    openai_api_base=st.session_state.base_url,
                                    temperature=0.3,
                                )
                                segment_report_template = ChatPromptTemplate.from_template(
                                    """
你是一名专业的游戏用户研究员。你具备极强的文本洞察力，能够准确归因并提炼玩家真实诉求。
本次分析的核心研究目标是：{research_goal}。请仅针对指定的玩家分群撰写一份「深挖报告」。

目标玩家分群: {segment_name}

【数据说明】:
{segment_sample_note}

【该分群的样本结构化明细】(JSON 格式 samples_json):
{segment_samples_json}

【该分群的高频关键词统计】(keyword_context，JSON 格式):
{segment_keyword_json}

【该分群采样的原始玩家原话】(raw_quotes，JSON 格式):
{segment_raw_quotes_json}

写作要求:
1. 用 1 段话概括该分群在本次研究目标下的核心轮廓（他们是谁、处于什么情境、典型心智）。
2. 从「核心诉求关键词」中抽取 3-5 个最关键的需求，每一个需求都要有准确且客观实际的量化统计（比如有多少人提及，占比多少等），并用策划/产品视角解释它们具体意味着什么。
3. 结合 qualitative_insight，总结该分群的典型体验维度与偏好特征（例如更关注节奏、反馈、难度、公平性等），而不是局限在“职业/外形”这种具体题材。
4. 分析该分群的情绪结构：正面驱动力是什么？负面吐槽集中在哪些体验点？这些情绪与关键词统计之间是否一致。
5. 给出 3-5 条专门面向该分群的设计/优化建议（可以是系统、内容、运营层面的），并尽量引用关键词统计或原话中的线索进行佐证。

输出格式:
- 使用中文分段撰写，不要输出 JSON。
- 在每一条关键洞察或设计建议下方，请从当前分组的原话池中选取 3-5 条最具代表性的玩家原话，以 Markdown 引用块（> 开头）或列表的形式展示，使报告更具现场感；
- 在分析时，如果你的语义结论与 keyword_context 中的频次一致，请在文字中引用相应数字进行佐证；
- 如果语义结论与 keyword_context 中的统计出现明显冲突，请在报告中指出可能原因（例如样本量不足、问卷题干歧义等），避免给出与统计显著不一致的定性判断。
                                    """
                                )
                                seg_messages = segment_report_template.format_messages(
                                    segment_name=f"{core_segment_col} = {selected_norm}",
                                    segment_sample_note=seg_sample_note,
                                    segment_keyword_json=json.dumps(segment_keyword_stats, ensure_ascii=False),
                                    segment_raw_quotes_json=json.dumps(segment_raw_quotes, ensure_ascii=False),
                                    segment_samples_json=json.dumps(seg_sample_records, ensure_ascii=False),
                                    research_goal=st.session_state.research_goal,
                                )
                                seg_response = llm_for_segment.invoke(seg_messages)
                                seg_report_text = seg_response.content
                                reports[selected_norm] = seg_report_text
                        except Exception as e:
                            reports[str(selected_segment)] = f"生成该分群深挖报告时出错: {e}"
                        progress_seg.progress(idx_seg / total_seg)
                        status_seg.text(f"正在生成第 {idx_seg}/{total_seg} 个分群报告：{selected_segment}")
                    st.session_state.segment_reports = reports
                    status_seg.text("所有选定分群的深挖报告已生成。")

        if "segment_reports" in st.session_state and st.session_state.segment_reports:
            reports = st.session_state.segment_reports
            names = sorted(reports.keys())
            tabs = st.tabs([f"{core_segment_col} = {name}" for name in names])
            for tab, name in zip(tabs, names):
                with tab:
                    st.text_area(f"分组 [{core_segment_col} = {name}] 深挖报告", reports[name], height=260)

    # --- 5.2 PPT 大纲生成 ---
    st.divider()
    st.subheader("📑 给策划看的 PPT 大纲")
    st.markdown("基于深度研究报告，自动生成适合策划快速过会的 PPT 提纲。")

    if st.button("生成 PPT 大纲"):
        base_report = st.session_state.get("deep_report", "")
        if not base_report:
            st.warning("请先生成整体深度报告。")
        else:
            try:
                llm_for_ppt = ChatOpenAI(
                    model=model_name,
                    openai_api_key=st.session_state.api_key,
                    openai_api_base=st.session_state.base_url,
                    temperature=0.2,
                )

                ppt_template = ChatPromptTemplate.from_template(
                    """
下面是一份针对特定研究目标的「深度研究报告」正文。当前研究目标为：{research_goal}。
请你将其整理成一份适合策划/制作人快速浏览的 PPT 大纲，使大纲结构和页标题与该研究目标高度相关。

【研究报告正文】:
{full_report}

输出要求:
1. 使用纯文本输出 PPT 大纲，不要使用任何 JSON 或 Markdown 代码块。
2. 结构建议（可在此基础上适度调整）:
   - 封面：项目名 + 研究主题
   - 目录
   - 研究背景与样本说明
   - 玩家分群与总体洞察
   - 核心体验维度与典型主题拆解
   - 情绪与体验风险
   - 可落地设计建议
3. 每一页使用格式：
   第 X 页：<页面标题>
   - 要点 1
   - 要点 2
   - 要点 3
4. 要点要尽量短句、可执行，方便直接拷贝到 PPT 中使用。
                    """
                )

                ppt_messages = ppt_template.format_messages(
                    full_report=base_report,
                    research_goal=st.session_state.research_goal,
                )
                ppt_response = llm_for_ppt.invoke(ppt_messages)
                ppt_outline_text = ppt_response.content
                st.session_state.ppt_outline = ppt_outline_text
                st.success("PPT 大纲已生成。")
            except Exception as e:
                st.error(f"生成 PPT 大纲时出错: {e}")

    if "ppt_outline" in st.session_state and st.session_state.ppt_outline:
        st.markdown("#### PPT 大纲预览")
        st.text_area("PPT 大纲", st.session_state.ppt_outline, height=260)

    # --- 6. 导出功能 ---
    st.divider()
    if st.button("🧰 准备导出文件"):
        with st.spinner("正在生成导出文件..."):
            report_text = st.session_state.get("deep_report", "")
            ppt_text = st.session_state.get("ppt_outline", "")
            segment_reports = st.session_state.get("segment_reports", {})
            export_data = build_export_workbook_bytes(
                adf=adf,
                deep_report_text=report_text,
                ppt_text=ppt_text,
                segment_reports=segment_reports,
            )
            st.session_state.export_data_bytes = export_data
            st.session_state.export_data_name = "analyzed_game_data_with_report_and_ppt.xlsx"
            st.session_state.export_data_seq = st.session_state.get("export_data_seq", 0) + 1
            st.success("导出文件已生成，可点击下方按钮下载。")

    if st.session_state.get("export_data_bytes"):
        _ex_suggested = st.session_state.get(
            "export_data_name", "analyzed_game_data_with_report_and_ppt.xlsx"
        )
        _ex_seq = st.session_state.get("export_data_seq", 0)
        if _ex_seq != st.session_state.get("export_data_fn_bound_seq", -1):
            st.session_state["export_data_dl_filename"] = _ex_suggested
            st.session_state["export_data_fn_bound_seq"] = _ex_seq
        st.text_input("下载文件名（可修改）", key="export_data_dl_filename")
        _ex_fn = safe_download_filename(
            st.session_state.get("export_data_dl_filename", _ex_suggested),
            fallback=_ex_suggested,
        )
        st.download_button(
            label="📥 导出明细、报告与 PPT 大纲 (Excel)",
            data=st.session_state.export_data_bytes,
            file_name=_ex_fn,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# --- 7. 无数据时的提示 ---
if uploaded_file is None:
    st.info("请在上方上传数据文件以开始分析。")
