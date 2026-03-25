# -*- coding: utf-8 -*-
"""
Playtest 自动化分析流水线 v0.3

维护约定（开发 / AI 必读）：
  每次变更 Pipeline 行为、CLI 参数或对外统计口径时，须同步更新：
    - docs/PLAYTEST_PIPELINE.md（完整说明与参数表）
    - 本模块 docstring（保持与上述文档一致）
    - docs/DEV_LOG.md（一条变更记录，便于审计）
  参数以 `python scripts/run_playtest_pipeline.py --help` 为准。

功能：
  1. 从 data/raw/ 自动读取最新问卷数据（.sav / .csv / .xlsx）
     [A] .sav 文件时打印格式提示，建议优先使用 .xlsx
  2. 自动识别题型
     [B] .sav 文件时利用 pyreadstat meta（variable_measure + 原始变量名前缀）辅助识别多选题
     [NEW] 利用列名 `N.题目—子题目` 模式自动识别矩阵题子项分组
  3. 寻找分组列
     [C] --segment-col 参数（子串模糊匹配）优先于自动识别
     [E] 无分组列时自动生成「总体」虚拟列，输出全量频率分布而非跳过
  4. 满意度回归：总样本量 N<15 跳过；15<=N<50 可跑回归并标记低样本预警（导出 A1 免责+浅灰表）；N>=50 正常
  5. 汇总导出为多 Sheet Excel
     [D] 按 Q 题号分组的独立题 Sheet（--per-question-sheets 开启时）
     [NEW] 矩阵题合并为 2D 表（行=子题目，列=量表选项，格=n(X.X%)）
     [NEW] 单选/评分题输出：选项|小计(n)|总体%，附均值行和有效填写人次
     [NEW] 利用大纲（.docx 问卷星 / .txt 腾讯问卷）补全 0% 选项、修正题型和排序
  6. （可选）显著性检验收口到 quant 引擎：评分题(k=2) Welch t、评分题(k>2) ANOVA/KW；单/多选卡方。

CLI 参数摘要：
  --data-dir, --output-dir, --segment-col, --sheet-name, --per-question-sheets, --outline,
  --sig-test / --no-sig-test, --sig-alpha
  详见 docs/PLAYTEST_PIPELINE.md 或 --help。

运行方式（在 UserResearch/ 目录下执行）：

    # 默认：自动读取 data/raw/ 最新文件
    python scripts/run_playtest_pipeline.py

    # 指定问卷大纲（.docx 问卷星 或 .txt 腾讯问卷，否则自动发现 data/raw/ 下最新文件）
    python scripts/run_playtest_pipeline.py --outline data/raw/questionnaire.docx
    python scripts/run_playtest_pipeline.py --outline data/raw/questionnaire.txt

    # 指定分组列（子串即可，如"玩家类型"，不需要输入完整列名）
    python scripts/run_playtest_pipeline.py --segment-col "玩家类型"

    # 关闭显著性检验 / 自定义阈值
    python scripts/run_playtest_pipeline.py --no-sig-test
    python scripts/run_playtest_pipeline.py --sig-alpha 0.01

    # 查看全部参数（与实现同步）
    python scripts/run_playtest_pipeline.py --help
"""
from __future__ import annotations

import datetime
import io as _io
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# Windows PowerShell 默认 GBK 不支持部分 Unicode 字符，强制 stdout/stderr 使用 UTF-8
if sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# 确保从 UserResearch/ 目录运行时可以找到 survey_tools 包
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import click
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from survey_tools.core.quant import extract_qnum
from survey_tools.core.quant import (
    build_question_specs,
    run_quant_cross_engine,
)
from survey_tools.core.question_type import (
    detect_column_type,
    get_prefix,
    infer_type_from_columns,
    parse_columns_for_questions,
)
from survey_tools.utils.io import (
    ExportBundle,
    export_xlsx,
    get_latest_local_data,
    load_survey_data,
)
from survey_tools.utils.outline_parser import parse_outline_docx, parse_outline_txt

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
# 满意度回归：总样本量 N=len(df) 分档（与 _run_satisfaction_modeling 一致）
_MIN_N_REGRESSION_SKIP = 15  # N < 此值：不执行回归
_MIN_N_REGRESSION_FULL = 50  # N >= 此值：无低样本预警；[15, 50) 为低样本预警档

# 满意度回归结果 Sheet：15<=N<50 时 A1 免责声明（{n} 为 low_sample_n，即总行数 len(df)）
_SATISFACTION_LOW_N_DISCLAIMER = (
    "【专业预警】当前分析样本量为 {n}，小于统计学建议的样本量下限 (50)。"
    "本回归模型可能存在过拟合风险，计算权重仅代表当前测试组的趋势波动，"
    "严禁将其作为整体大盘的定论，请结合定性反馈审慎参考。"
)
_FILL_SATISFACTION_LOW_N = PatternFill(
    start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
)
_FONT_SATISFACTION_WARN = Font(bold=True, color="FF0000")
_MAX_INDIVIDUAL_SHEETS = 60
_SEGMENT_KEYWORDS = [
    "组别", "分组", "类型", "分类", "层级", "经验", "玩家类型", "group", "type", "segment",
]
_SATISFY_KEYWORDS = [
    "总体满意", "满意度", "nps", "综合评分", "整体评分", "总满意", "总评", "overall",
]
_NPS_KEYWORDS = [
    "nps", "净推荐", "推荐意愿", "推荐可能性", "推荐概率", "recommend",
]
_FORCE_IGNORE_KEYWORDS = [
    "答卷时间", "ip", "所用时间", "序号", "总分", "答卷编号", "逻辑", "跳转",
]
_TYPE_RANK: Dict[str, int] = {
    "单选": 1, "评分": 2, "多选": 3, "矩阵单选": 4, "矩阵评分": 5, "矩阵": 6,
}
_SYNTHETIC_SEGMENT_COL = "_总体_"

# 矩阵题子项分隔符（问卷星导出格式：题干—子题目）
_MATRIX_SEP = "—"


# ---------------------------------------------------------------------------
# 辅助：从选项文本中提取数值分数（如 "4分（比较满意）" → 4.0）
# ---------------------------------------------------------------------------
def _extract_option_value(opt_text: str) -> Optional[float]:
    """从选项标签提取开头的数字分值，用于计算均值。

    Args:
        opt_text: 选项文本，如 "4分（比较满意）"、"5"、"非常同意"。

    Returns:
        float 分值，无法提取时返回 None。
    """
    m = re.match(r"^\s*(\d+(?:\.\d+)?)", str(opt_text).strip())
    return float(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# 大纲：自动发现 data/raw/ 中最新的 .docx / .txt 大纲文件
# ---------------------------------------------------------------------------
def _get_latest_local_outline(folder_path: str | Path) -> Optional[Path]:
    """扫描目录，找出最近修改的大纲文件（.docx 或 .txt）。

    Args:
        folder_path: 扫描目录路径。

    Returns:
        最新大纲文件的 Path，目录不存在或无候选文件时返回 None。
    """
    folder = Path(folder_path)
    if not folder.exists():
        return None
    candidates = list(folder.glob("*.docx")) + list(folder.glob("*.txt"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


# ---------------------------------------------------------------------------
# 矩阵题子项分组：利用列名 `N.题干—子题目` 模式识别矩阵列
# ---------------------------------------------------------------------------
def _infer_matrix_groups(columns: List[str]) -> Dict[str, int]:
    """根据列名模式 `N.题干—子题目` 推断矩阵题子项与题号的映射。

    问卷星矩阵题导出格式：
      - 第一子项：`N.题干...—子题目文本`（含题号 + _MATRIX_SEP 分隔符）
      - 后续子项：`子题目文本`（无题号，无分隔符）

    此函数通过 `_MATRIX_SEP` 的出现标记第一子项，并将随后不含题号的列
    视为同一矩阵题的子项，直到遇到下一个带题号的列为止。

    Args:
        columns: DataFrame 列名列表。

    Returns:
        dict，列名 → 所属矩阵题题号。仅矩阵题子项出现在结果中。
    """
    matrix_map: Dict[str, int] = {}
    current_matrix_q: Optional[int] = None
    q_num_pattern = re.compile(r"^(\d+)\.")

    for col in columns:
        col_s = str(col)
        m = q_num_pattern.match(col_s)
        if m:
            q_num = int(m.group(1))
            if _MATRIX_SEP in col_s:
                # 矩阵题第一子项
                current_matrix_q = q_num
                matrix_map[col_s] = q_num
            else:
                # 新的带题号列，且不含分隔符 → 非矩阵子项，重置状态
                current_matrix_q = None
        else:
            # 无题号列：若当前处于矩阵题跟踪中，视为子项
            if current_matrix_q is not None:
                matrix_map[col_s] = current_matrix_q

    return matrix_map


# ---------------------------------------------------------------------------
# Step 1：数据装载
# ---------------------------------------------------------------------------
def _load_data(
    data_dir: str,
    sheet_name: Optional[int | str] = 0,
) -> Tuple[pd.DataFrame, Optional[Any]]:
    """读取最新问卷数据，同时返回 .sav 的原始 meta 对象。

    Args:
        data_dir: 扫描目录路径。

    Returns:
        (df, sav_meta)，sav_meta 仅 .sav 文件时有值。
    """
    print(f"\n[1/4] 数据装载 ...")
    path = get_latest_local_data(data_dir)
    print(f"  → 找到最新数据文件：{path}")

    sav_meta: Optional[Any] = None
    suffix = Path(str(path)).suffix.lower()

    if suffix == ".sav":
        print("  ⚠  [格式提示] 检测到 .sav 文件。")
        print("     .sav 的题型识别（尤其是多选题）依赖变量标签格式，可能不如标准")
        print("     问卷星 .xlsx 准确。如有 .xlsx 版本，建议优先使用 .xlsx。")
        try:
            import pyreadstat as _prs
            _, sav_meta = _prs.read_sav(str(path))
        except Exception as _e:
            print(f"     （meta 加载失败，将跳过 meta 增强识别：{_e}）")

    df = load_survey_data(path, sheet_name=sheet_name if sheet_name is not None else 0)
    if str(path).lower().endswith((".xlsx", ".xls")) and sheet_name is not None:
        print(f"  → Excel 读取 Sheet：{sheet_name}")
    print(f"  → 成功读取数据，样本量：{len(df)}，列数：{len(df.columns)}")
    return df, sav_meta


# ---------------------------------------------------------------------------
# 辅助 B：利用 .sav meta 增强多选题识别
# ---------------------------------------------------------------------------
def _enhance_with_sav_meta(
    column_type_map: Dict[str, str],
    df: pd.DataFrame,
    sav_meta: Any,
) -> Dict[str, str]:
    """利用 .sav meta 信息将被误判为单选的多选子列修正为多选。

    策略：
    1. 原始变量名前缀分组（如 Q5_1/Q5_2/Q5_3 → 前缀 Q5）：若同一前缀下有 2+
       个 0/1 二分列，则归为多选。
    2. variable_measure == 'scale' 的列若当前被标注为单选，改为评分。

    Args:
        column_type_map: 当前题型映射（将被修改并返回）。
        df: 已应用变量标签的 DataFrame（列名为题目文本）。
        sav_meta: pyreadstat.read_sav 返回的 meta 对象。

    Returns:
        修正后的 column_type_map。
    """
    result = dict(column_type_map)

    orig_to_label: Dict[str, str] = dict(sav_meta.column_names_to_labels or {})
    label_to_orig: Dict[str, str] = {v: k for k, v in orig_to_label.items()}
    var_measures: Dict[str, str] = getattr(sav_meta, "variable_measure", {}) or {}

    prefix_groups: Dict[str, List[str]] = defaultdict(list)
    for orig in (sav_meta.column_names or []):
        stripped = re.sub(r"[_\s]?\d+$", "", orig)
        prefix_groups[stripped].append(orig)

    for _prefix, orig_list in prefix_groups.items():
        if len(orig_list) < 2:
            continue
        label_cols = [orig_to_label.get(o) for o in orig_list]
        label_cols = [c for c in label_cols if c and c in df.columns]
        if len(label_cols) < 2:
            continue
        all_binary = all(
            set(
                float(x)
                for x in pd.to_numeric(df[lc], errors="coerce").dropna().unique()
            ).issubset({0.0, 1.0})
            for lc in label_cols
        )
        if all_binary:
            for lc in label_cols:
                result[lc] = "多选"

    for col in df.columns:
        orig = label_to_orig.get(col)
        if orig and var_measures.get(orig) == "scale":
            if result.get(col) == "单选":
                result[col] = "评分"

    return result


# ---------------------------------------------------------------------------
# 辅助：自动题型识别（含大纲覆盖 + 矩阵子项识别 + B 增强入口）
# ---------------------------------------------------------------------------
def _auto_classify_columns(
    df: pd.DataFrame,
    sav_meta: Optional[Any] = None,
    outline: Optional[Dict[int, dict]] = None,
    matrix_q_map: Optional[Dict[str, int]] = None,
) -> Dict[str, str]:
    """自动推断每列的题型，返回 column_type_map（列名→题型）。

    优先级：大纲覆盖 > 矩阵子项模式识别 > 列名关键词分组推断 > detect_column_type。

    Args:
        df: 原始问卷 DataFrame。
        sav_meta: pyreadstat meta 对象（.sav 文件时传入，否则 None）。
        outline: _parse_outline_docx 返回的大纲信息（可为 None）。
        matrix_q_map: _infer_matrix_groups 返回的矩阵列→题号映射（可为 None）。

    Returns:
        dict，列名 → 题型字符串。
    """
    columns = df.columns.tolist()
    questions_data = parse_columns_for_questions(columns)

    q_num_to_type: Dict[int, str] = {}
    for q_num, info in questions_data.items():
        raw = infer_type_from_columns(info)
        if raw:
            if "矩阵单选" in raw:
                q_num_to_type[q_num] = "矩阵单选"
            elif "矩阵评分" in raw:
                q_num_to_type[q_num] = "矩阵评分"
            elif "矩阵" in raw:
                q_num_to_type[q_num] = "矩阵单选"
            elif "多选" in raw:
                q_num_to_type[q_num] = "多选"
            elif "评分" in raw:
                q_num_to_type[q_num] = "评分"
            elif "填空" in raw or "开放" in raw:
                q_num_to_type[q_num] = "忽略"
            else:
                q_num_to_type[q_num] = "单选"

    # 大纲覆盖：仅覆盖大纲可靠区分的题型（矩阵/多选/填空）
    # 注意：大纲里量表题（1-5 分）也标为"单选题"，无法区分单选 vs 评分，
    # 因此"单选题"类型不做覆盖，保留数据驱动识别（以免把"评分"列误标为"单选"导致回归失效）。
    if outline:
        for q_num, info in outline.items():
            otype = info["type"]
            if "矩阵评分" in otype:
                q_num_to_type[q_num] = "矩阵评分"
            elif "矩阵" in otype:
                # 矩阵文本题（填空式矩阵）→ 忽略；其余矩阵题 → 矩阵单选
                if "文本" in otype or "填空" in otype:
                    q_num_to_type[q_num] = "忽略"
                else:
                    q_num_to_type[q_num] = "矩阵单选"
            elif "多选" in otype:
                q_num_to_type[q_num] = "多选"
            elif "填空" in otype or "文本" in otype:
                q_num_to_type[q_num] = "忽略"
            # "单选题" / "评分题" 不覆盖，由数据值特征决定

    known_multi_prefixes: set = set()
    for col in columns:
        q_str = extract_qnum(str(col))
        if q_str:
            try:
                if q_num_to_type.get(int(q_str)) == "多选":
                    known_multi_prefixes.add(get_prefix(str(col)))
            except ValueError:
                pass

    column_type_map: Dict[str, str] = {}
    force_ignored_cols: List[str] = []
    for col in columns:
        col_s = str(col)
        # 强制忽略系统元数据列（中英文括号/大小写兼容）
        col_norm = (
            col_s.replace("【", "")
            .replace("】", "")
            .replace("[", "")
            .replace("]", "")
            .strip()
            .lower()
        )
        if any(k in col_norm for k in _FORCE_IGNORE_KEYWORDS):
            column_type_map[col] = "忽略"
            force_ignored_cols.append(col_s)
            continue

        # 腾讯问卷填空伴随列（如 "N.题干:其他__[选项填空]"）→ 直接忽略
        if "[选项填空]" in col_s:
            column_type_map[col] = "忽略"
            continue

        q_str = extract_qnum(col_s)
        q_num: Optional[int] = None
        if q_str:
            try:
                q_num = int(q_str)
            except ValueError:
                pass

        if q_num is not None and q_num in q_num_to_type:
            column_type_map[col] = q_num_to_type[q_num]
        elif matrix_q_map and col_s in matrix_q_map:
            # 矩阵子项（无题号列），从 matrix_q_map 取得所属题号的题型
            parent_q = matrix_q_map[col_s]
            column_type_map[col] = q_num_to_type.get(parent_q, "矩阵单选")
        else:
            column_type_map[col] = detect_column_type(
                col, df[col], get_prefix(col_s), known_multi_prefixes
            )

    # ── 腾讯问卷格式补充：将 "N.题干:子题" 格式的矩阵列加入 matrix_q_map ──────────
    # 腾讯问卷使用 ":" 分隔题干与子项，且所有子列均带 N. 前缀；
    # _infer_matrix_groups 只识别 "—" 格式，对腾讯列无效。
    # 此处补充：对 q_num_to_type 中类型为矩阵的题号，将其名下含 ":" 的子列记入
    # matrix_q_map（in-place 修改，后续 _run_quant_cross / _export_results 可见）。
    if matrix_q_map is not None:
        _MATRIX_TYPES = {"矩阵单选", "矩阵评分", "矩阵"}
        for col in columns:
            col_s = str(col)
            # 仅处理含半角冒号、不含 em-dash、且尚未在 matrix_q_map 中的列
            if ":" in col_s and _MATRIX_SEP not in col_s and col_s not in matrix_q_map:
                q_str2 = extract_qnum(col_s)
                if q_str2:
                    try:
                        q_num_c = int(q_str2)
                        if q_num_to_type.get(q_num_c) in _MATRIX_TYPES:
                            matrix_q_map[col_s] = q_num_c
                    except ValueError:
                        pass

    if sav_meta is not None:
        try:
            column_type_map = _enhance_with_sav_meta(column_type_map, df, sav_meta)
        except Exception as _e:
            print(f"  ⚠  meta 增强识别失败（已跳过）：{_e}")

    # 运行前打印「强制忽略」列清单，便于核对排除范围
    print(f"  → 强制忽略列（{len(force_ignored_cols)}）：")
    for c in force_ignored_cols:
        print(f"    - {c}")

    return column_type_map


# ---------------------------------------------------------------------------
# 辅助：寻找分组列（含方向 C 显式参数 + 方向 E 总体 fallback）
# ---------------------------------------------------------------------------
def _resolve_segment_col(
    df: pd.DataFrame,
    column_type_map: Dict[str, str],
    segment_col_hint: Optional[str],
) -> Tuple[pd.DataFrame, str, bool]:
    """确定最终使用的分组列，按优先级：显式参数 > 自动关键词识别 > 「总体」fallback。

    Args:
        df: 问卷 DataFrame（可能被修改添加 _总体_ 列）。
        column_type_map: 列名→题型映射。
        segment_col_hint: 来自 --segment-col 参数的子串（可为 None）。

    Returns:
        (df, segment_col, is_synthetic)
    """
    if segment_col_hint:
        matched = [c for c in df.columns if segment_col_hint in str(c)]
        if matched:
            print(f"  → [--segment-col] 匹配到分组列：「{matched[0]}」")
            return df, matched[0], False
        print(f"  ⚠  --segment-col 未匹配到含「{segment_col_hint}」的列，回退到自动识别。")

    for col in df.columns:
        if column_type_map.get(col) in ("忽略", "评分", "矩阵评分"):
            continue
        lower = str(col).lower()
        for kw in _SEGMENT_KEYWORDS:
            if kw in lower:
                n_unique = df[col].nunique()
                if 2 <= n_unique <= 20:
                    print(f"  → 自动识别分组列：「{col}」（{n_unique} 组）")
                    return df, col, False

    print("  → 未找到分组列，使用「总体」作为单一分组，输出全量频率分布。")
    df = df.copy()
    df[_SYNTHETIC_SEGMENT_COL] = "总体"
    return df, _SYNTHETIC_SEGMENT_COL, True


# ---------------------------------------------------------------------------
# 辅助：单选/评分题透视 + 选项补全（reindex）
# ---------------------------------------------------------------------------
def _simple_pivot(res: dict, option_list: Optional[List[str]] = None) -> pd.DataFrame:
    """将 run_quant_cross_engine 长表透视为宽表（分组→列），并可按大纲选项补全。

    输出列顺序：选项 | 小计(n) | 总体% | [group1] | [group2] | ...
    若 option_list 不为空，按大纲顺序排列选项，未出现的选项补 0。

    Args:
        res: run_quant_cross_engine 返回列表中的单个元素 dict。
        option_list: 大纲中的完整选项列表（用于 reindex，可为 None）。

    Returns:
        透视后的 DataFrame；若失败则返回原始长表。
    """
    df_q = res["数据"]
    q_type = res["题型"]
    if df_q.empty:
        return df_q
    try:
        if q_type in ("单选", "评分", "矩阵单选", "矩阵评分"):
            # Step 1: 透视（保留原始选项值）
            pivot = df_q.pivot_table(
                index="选项",
                columns="核心分组",
                values="行百分比",
                aggfunc="first",
                fill_value=0,
            ).reset_index()

            # Step 2: 计算总计（用原始值作为 key）
            total_counts = df_q.groupby("选项")["频次"].sum()
            total_samples = df_q["频次"].sum()
            total_ratio = total_counts / total_samples if total_samples > 0 else 0

            pivot.insert(1, "总体%", pivot["选项"].map(total_ratio).fillna(0))

            # Step 3: 格式化百分比列（总体% 和各分组列，不含选项/人数N）
            pct_cols = [c for c in pivot.columns if c != "选项"]
            for c in pct_cols:
                pivot[c] = pivot[c].apply(
                    lambda v: f"{v:.1%}"
                    if isinstance(v, (float, int)) and not isinstance(v, bool)
                    else v
                )

            # Step 4: 去掉与"总体%"重复的合成分组列
            # 合成分组时 df["_总体_"] = "总体"，pivot_table 会产生名为"总体"的列
            # 当且仅当唯一的分组列恰好叫"总体"时才删除，避免误删真实组名
            _non_stat_cols = {"选项", "总体%"}
            group_cols_in_pivot = [c for c in pivot.columns if c not in _non_stat_cols]
            if len(group_cols_in_pivot) == 1 and group_cols_in_pivot[0] == "总体":
                pivot = pivot.drop(columns=["总体"])

            # Step 5: 在末尾追加人数N列（基于原始值，在选项重命名前计算映射）
            n_map = pivot["选项"].map(total_counts).fillna(0)

            # Step 6: 大纲选项重命名 + 零填充（在统计计算完成后进行）
            # 问卷星 Excel 存储的是数值（如整数 4），大纲选项是文本（如"4分（比较满意）"）
            if option_list:
                # 构建 原始值字符串 → 大纲标签文本 的映射
                val_to_label: Dict[str, str] = {}
                for opt in option_list:
                    v = _extract_option_value(str(opt))
                    if v is not None:
                        val_to_label[str(int(v))] = opt    # "4" → "4分（比较满意）"
                        val_to_label[str(v)] = opt         # "4.0" → "4分（比较满意）"

                # 如果数字前缀映射与实际数据完全不重叠（例如选项文本是"50小时以下"
                # 提取出50，但数据值是1/2/3/4/5），则自动降级为按位置映射
                def _build_pos_map(opts: List[str]) -> Dict[str, str]:
                    m: Dict[str, str] = {}
                    for i, o in enumerate(opts):
                        m[str(i + 1)] = o
                        m[f"{i + 1}.0"] = o   # float 字符串变体
                    return m

                if val_to_label:
                    actual_vals = {str(x) for x in pivot["选项"]}
                    if not (actual_vals & set(val_to_label.keys())):
                        # 数字前缀全部失配，尝试位置映射
                        pos_map = _build_pos_map(option_list)
                        if actual_vals & set(pos_map.keys()):
                            val_to_label = pos_map
                else:
                    # 无数字前缀时按位置映射（值"1"→第1个选项，值"2"→第2个选项，…）
                    val_to_label = _build_pos_map(option_list)

                # 将 "选项" 列原始值替换为大纲标签
                if val_to_label:
                    pivot["选项"] = pivot["选项"].apply(
                        lambda x: val_to_label.get(str(x), str(x))
                    )

                # 补全零频次选项（出现在大纲但数据中没有的选项）
                existing = set(pivot["选项"].tolist())
                zero_pct_fmt = "0.0%"
                for opt in option_list:
                    if opt not in existing:
                        zero_row: Dict[str, Any] = {c: zero_pct_fmt for c in pivot.columns}
                        zero_row["选项"] = opt
                        n_map = pd.concat([n_map, pd.Series([0])], ignore_index=True)
                        pivot = pd.concat(
                            [pivot, pd.DataFrame([zero_row])], ignore_index=True
                        )

                # 按大纲选项顺序重排（未在大纲中的附在末尾）
                opt_order = {opt: i for i, opt in enumerate(option_list)}
                sort_idx = pivot["选项"].map(lambda x: opt_order.get(x, len(option_list)))
                pivot = pivot.iloc[sort_idx.argsort().values].reset_index(drop=True)
                n_map = n_map.iloc[sort_idx.argsort().values].reset_index(drop=True)

            pivot["人数N"] = n_map.fillna(0).astype(int)
        else:
            # 多选题 — 与单选格式一致：选项 | 总体% | [group%] | 人数N
            pivot = df_q.pivot_table(
                index="选项",
                columns="核心分组",
                values="提及率",
                aggfunc="first",
                fill_value=0,
            ).reset_index()
            total_mentions = df_q.groupby("选项")["提及人数"].sum()
            total_sample = df_q.groupby("核心分组")["组样本数"].first().sum()
            overall_rate = total_mentions / total_sample if total_sample > 0 else 0
            pivot.insert(1, "总体%", pivot["选项"].map(overall_rate))

            # 去掉与"总体%"重复的合成分组列（与单选一致）
            _non_stat_cols = {"选项", "总体%"}
            group_cols_in_pivot = [c for c in pivot.columns if c not in _non_stat_cols]
            if len(group_cols_in_pivot) == 1 and group_cols_in_pivot[0] == "总体":
                pivot = pivot.drop(columns=["总体"])

            # 格式化百分比列
            pct_cols = [c for c in pivot.columns if c != "选项"]
            for c in pct_cols:
                pivot[c] = pivot[c].apply(
                    lambda v: f"{v:.1%}"
                    if isinstance(v, (float, int)) and not isinstance(v, bool)
                    else v
                )

            # 人数N：提及该选项的人数（跨组求和）
            n_map = pivot["选项"].map(total_mentions).fillna(0)

            # 大纲选项补全 0% 与排序（与单选一致）
            if option_list:
                existing = set(pivot["选项"].tolist())
                zero_pct_fmt = "0.0%"
                for opt in option_list:
                    if opt not in existing:
                        zero_row: Dict[str, Any] = {c: zero_pct_fmt for c in pivot.columns}
                        zero_row["选项"] = opt
                        zero_row["总体%"] = zero_pct_fmt
                        pivot = pd.concat(
                            [pivot, pd.DataFrame([zero_row])], ignore_index=True
                        )
                        n_map = pd.concat([n_map, pd.Series([0])], ignore_index=True)
                # 按大纲选项顺序重排
                opt_order = {opt: i for i, opt in enumerate(option_list)}
                sort_idx = pivot["选项"].map(lambda x: opt_order.get(x, len(option_list)))
                pivot = pivot.iloc[sort_idx.argsort().values].reset_index(drop=True)
                n_map = n_map.iloc[sort_idx.argsort().values].reset_index(drop=True)

            pivot["人数N"] = n_map.fillna(0).astype(int)
        return pivot
    except Exception:
        return df_q


# ---------------------------------------------------------------------------
# 辅助：单道非矩阵题的完整格式化 Block（含均值行 + N 行）
# ---------------------------------------------------------------------------
def _build_question_block(
    res: dict,
    option_list: Optional[List[str]] = None,
) -> pd.DataFrame:
    """构建单道非矩阵题的完整格式化 DataFrame block，用于汇总 Sheet 纵向拼接。

    格式：
      [题目标题行]
      [本题平均分行]（仅数值选项题）
      [Banner N 行]（样本量(N) | 总体N | 各组N）
      [列头行]（选项 | 总体% | ...）
      [pivot 数据行]
      [空行]

    Args:
        res: run_quant_cross_engine 结果 dict。
        option_list: 大纲选项列表，传给 _simple_pivot 补全 0 值。

    Returns:
        DataFrame，所有列与 pivot_df 对齐。
    """
    question = str(res.get("题目", ""))
    q_type = res.get("题型", "")
    df_q = res["数据"]

    pivot_df = _simple_pivot(res, option_list=option_list)
    if pivot_df.empty:
        return pivot_df

    col0 = pivot_df.columns[0]  # "选项"

    def _make_row(label: str) -> pd.DataFrame:
        row = {c: "" for c in pivot_df.columns}
        row[col0] = label
        return pd.DataFrame([row])

    def _is_nps_question(question_text: str, data: pd.DataFrame) -> bool:
        q = str(question_text).strip().lower()
        has_nps_keyword = any(k in q for k in _NPS_KEYWORDS)
        # 兼容常见中文问法：同时出现“推荐”与“意愿/可能/多大”
        has_cn_pattern = ("推荐" in q) and any(k in q for k in ("意愿", "可能", "多大"))
        if not (has_nps_keyword or has_cn_pattern):
            return False
        opt_vals = [_extract_option_value(str(opt)) for opt in data.get("选项", pd.Series(dtype=object)).unique()]
        numeric_vals = [v for v in opt_vals if v is not None]
        if not numeric_vals:
            return False
        return min(numeric_vals) >= 0 and max(numeric_vals) <= 10

    def _calc_nps_rows(data: pd.DataFrame) -> List[pd.DataFrame]:
        rows: List[pd.DataFrame] = []
        if data.empty:
            return rows

        by_opt = data.groupby("选项")["频次"].sum()
        to_score = {
            opt: _extract_option_value(str(opt))
            for opt in by_opt.index
        }

        promoter_opts = [opt for opt, v in to_score.items() if v is not None and v >= 9]
        passive_opts = [opt for opt, v in to_score.items() if v is not None and 7 <= v <= 8]
        detractor_opts = [opt for opt, v in to_score.items() if v is not None and v <= 6]

        grp_n_local = data.groupby("核心分组")["组样本数"].first()
        total_n_local = int(grp_n_local.sum())

        def _rate_for(opts: List[Any]) -> tuple[float, Dict[str, float]]:
            total_cnt = int(data["频次"].sum())
            overall_rate = (data[data["选项"].isin(opts)]["频次"].sum() / total_cnt) if total_cnt > 0 else 0.0
            per_group_rate: Dict[str, float] = {}
            for grp in data["核心分组"].dropna().unique():
                gdf = data[data["核心分组"] == grp]
                g_n = int(gdf["组样本数"].iloc[0]) if not gdf.empty else 0
                g_cnt = int(gdf[gdf["选项"].isin(opts)]["频次"].sum())
                per_group_rate[str(grp)] = (g_cnt / g_n) if g_n > 0 else 0.0
            return overall_rate, per_group_rate

        promoter_overall, promoter_grp = _rate_for(promoter_opts)
        passive_overall, passive_grp = _rate_for(passive_opts)
        detractor_overall, detractor_grp = _rate_for(detractor_opts)

        def _fill_pct_row(label: str, overall: float, grp_map: Dict[str, float]) -> pd.DataFrame:
            r = _make_row(label)
            for j, c in enumerate(pivot_df.columns):
                if c == "选项":
                    continue
                if c == "总体%":
                    r.iloc[0, j] = f"{overall:.1%}"
                elif c in grp_map:
                    r.iloc[0, j] = f"{grp_map[c]:.1%}"
            return r

        rows.append(_fill_pct_row("Promoter占比（9-10）", promoter_overall, promoter_grp))
        rows.append(_fill_pct_row("Passive占比（7-8）", passive_overall, passive_grp))
        rows.append(_fill_pct_row("Detractor占比（0-6）", detractor_overall, detractor_grp))

        nps_row = _make_row("NPS（%Promoter-%Detractor）")
        for j, c in enumerate(pivot_df.columns):
            if c == "选项":
                continue
            if c == "总体%":
                nps_row.iloc[0, j] = f"{(promoter_overall - detractor_overall) * 100:.1f}"
            elif c in promoter_grp and c in detractor_grp:
                nps_row.iloc[0, j] = f"{(promoter_grp[c] - detractor_grp[c]) * 100:.1f}"
            elif c == "人数N":
                nps_row.iloc[0, j] = total_n_local
            elif c in grp_n_local.index:
                nps_row.iloc[0, j] = int(grp_n_local[c])
        rows.append(nps_row)
        return rows

    is_nps = q_type in ("单选", "评分") and _is_nps_question(question, df_q)

    # 计算均值（仅当选项含数字分值时；NPS 题不输出均值，避免与国际口径混淆）
    mean_val: Optional[float] = None
    group_means: Dict[str, float] = {}
    n_valid = 0
    if q_type in ("单选", "评分") and not df_q.empty and not is_nps:
        try:
            total_counts = df_q.groupby("选项")["频次"].sum()
            opt_vals = {opt: _extract_option_value(str(opt)) for opt in total_counts.index}
            numeric_opts = {opt: v for opt, v in opt_vals.items() if v is not None}
            if numeric_opts:
                weighted = sum(
                    numeric_opts[opt] * total_counts.get(opt, 0)
                    for opt in numeric_opts
                )
                n_valid = int(sum(total_counts.get(opt, 0) for opt in numeric_opts))
                if n_valid > 0:
                    mean_val = weighted / n_valid

                # 同时计算各组均值（用于导出时显示及显著性标记）
                for grp in df_q["核心分组"].dropna().unique():
                    gdf = df_q[df_q["核心分组"] == grp]
                    g_total_counts = gdf.groupby("选项")["频次"].sum()
                    g_weighted = sum(
                        numeric_opts.get(opt, 0.0) * g_total_counts.get(opt, 0)
                        for opt in g_total_counts.index
                        if opt in numeric_opts
                    )
                    g_n_valid = int(
                        sum(g_total_counts.get(opt, 0) for opt in g_total_counts.index if opt in numeric_opts)
                    )
                    if g_n_valid > 0:
                        group_means[str(grp)] = g_weighted / g_n_valid
            else:
                n_valid = int(total_counts.sum())
        except Exception:
            pass

    # Banner N 行：各组样本量，与 pivot 列对齐
    grp_n = df_q.groupby("核心分组")["组样本数"].first()
    total_n = int(grp_n.sum())
    banner_n_row = _make_row("")
    banner_n_row.iloc[0, 0] = "样本量(N)"
    for j, col in enumerate(pivot_df.columns):
        if col == "选项":
            continue
        if col == "总体%":
            banner_n_row.iloc[0, j] = total_n
        elif col == "人数N":
            banner_n_row.iloc[0, j] = ""  # 样本量行的人数N列留空
        elif col in grp_n.index:
            banner_n_row.iloc[0, j] = int(grp_n[col])

    # 列头行（选项|总体%|...|人数N）
    header_row = pd.DataFrame([dict(zip(pivot_df.columns, pivot_df.columns))])

    # T2B/B2B 行（仅 5 点量表：选项可解析出 4、5 分）
    t2b_b2b_rows: List[pd.DataFrame] = []
    if q_type in ("单选", "评分") and not df_q.empty and not is_nps:
        opt_vals = {}
        for opt in df_q["选项"].unique():
            v = _extract_option_value(str(opt))
            if v is not None:
                opt_vals[opt] = v
        has_4_5 = 4 in opt_vals.values() or 5 in opt_vals.values()
        if has_4_5:
            grp_n = df_q.groupby("核心分组")["组样本数"].first()
            total_n = int(grp_n.sum())
            opts_4_5 = [o for o, v in opt_vals.items() if v in (4, 5)]
            opts_1_2 = [o for o, v in opt_vals.items() if v in (1, 2)]
            pct_cols = [c for c in pivot_df.columns if c not in ("选项", "人数N")]

            def _calc_box(df_in: pd.DataFrame, opts: List[Any]) -> Tuple[float, Dict[str, float]]:
                total_cnt = df_in["频次"].sum()
                if total_cnt == 0:
                    return 0.0, {str(g): 0.0 for g in df_in["核心分组"].unique()}
                box_cnt = df_in[df_in["选项"].isin(opts)]["频次"].sum()
                overall = box_cnt / total_cnt
                per_grp = {}
                for g in df_in["核心分组"].unique():
                    gdf = df_in[df_in["核心分组"] == g]
                    gtot = gdf["组样本数"].iloc[0]
                    gbox = gdf[gdf["选项"].isin(opts)]["频次"].sum()
                    per_grp[str(g)] = gbox / gtot if gtot > 0 else 0.0
                return overall, per_grp

            t2b_overall, t2b_grp = _calc_box(df_q, opts_4_5)
            t2b_row = _make_row("T2B（4+5分）")
            for j, col in enumerate(pivot_df.columns):
                if col == "选项":
                    continue
                if col == "总体%":
                    t2b_row.iloc[0, j] = f"{t2b_overall:.1%}"
                elif col in t2b_grp:
                    t2b_row.iloc[0, j] = f"{t2b_grp[col]:.1%}"
            t2b_b2b_rows.append(t2b_row)

            if opts_1_2:
                b2b_overall, b2b_grp = _calc_box(df_q, opts_1_2)
                b2b_row = _make_row("B2B（1+2分）")
                for j, col in enumerate(pivot_df.columns):
                    if col == "选项":
                        continue
                    if col == "总体%":
                        b2b_row.iloc[0, j] = f"{b2b_overall:.1%}"
                    elif col in b2b_grp:
                        b2b_row.iloc[0, j] = f"{b2b_grp[col]:.1%}"
                t2b_b2b_rows.append(b2b_row)

    # NPS 国际口径：NPS = %Promoter(9-10) - %Detractor(0-6)
    nps_rows: List[pd.DataFrame] = []
    if is_nps:
        nps_rows = _calc_nps_rows(df_q)

    # 拼装 block：题目标题 → [均值] → Banner N → 列头 → 数据 → [T2B/B2B] → 空行
    parts: List[pd.DataFrame] = [
        _make_row(f"【{question}】（{q_type}）"),
    ]
    if mean_val is not None:
        mean_row = _make_row("本题平均分")
        for j, col in enumerate(pivot_df.columns):
            if col == "选项":
                mean_row.iloc[0, j] = "本题平均分"
            elif col == "总体%":
                mean_row.iloc[0, j] = f"{mean_val:.2f}"
            elif col == "人数N":
                mean_row.iloc[0, j] = ""
            elif str(col) in group_means:
                mean_row.iloc[0, j] = f"{group_means[str(col)]:.2f}"
        parts.append(mean_row)
    parts.append(banner_n_row)
    parts.append(header_row)
    parts.append(pivot_df)
    parts.extend(t2b_b2b_rows)
    parts.extend(nps_rows)
    parts.append(_make_row(""))

    return pd.concat(parts, ignore_index=True)


# ---------------------------------------------------------------------------
# 辅助：展开 "1~5" / "1-5" 类范围描述符为具体选项列表
# ---------------------------------------------------------------------------
def _expand_scale_options(options: Optional[List[str]]) -> Optional[List[str]]:
    """将形如 ["1~5"] 的量表范围描述符展开为 ["1","2","3","4","5"]。

    腾讯问卷矩阵量表题的量表行只写 "1~5"，需要展开才能正确映射数据值。
    若不是范围描述符则原样返回。

    Args:
        options: 大纲 options 列表（可为 None）。

    Returns:
        展开后的选项列表，或原始列表。
    """
    if not options or len(options) != 1:
        return options
    m = re.match(r"^(\d+)\s*[~～\-–]\s*(\d+)$", options[0].strip())
    if not m:
        return options
    lo, hi = int(m.group(1)), int(m.group(2))
    if 2 <= (hi - lo + 1) <= 15:   # 合理量表范围，最多15级
        return [str(i) for i in range(lo, hi + 1)]
    return options


# 矩阵量表题（评分矩阵）的大纲类型关键词
_RATING_OUTLINE_TYPES = {"矩阵量表题", "量表题"}


# ---------------------------------------------------------------------------
# 辅助：矩阵题 2D 汇总表
# ---------------------------------------------------------------------------
def _build_matrix_2d_table(
    group_results: List[dict],
    scale_options: Optional[List[str]] = None,
    is_rating: bool = False,
) -> pd.DataFrame:
    """将一道矩阵题的所有子项汇总为 2D 表。

    两种模式（由 is_rating 控制）：
      - 矩阵单选（is_rating=False）：行=子题目，列=量表选项，格=X.X%，末尾加人数N列
      - 矩阵评分（is_rating=True） ：行=子题目，列=分组，格=加权平均分

    Args:
        group_results: 同一矩阵题的所有子项结果 list（每项对应一个子题目）。
        scale_options: 大纲量表选项列表（矩阵单选用于列排序/补全；矩阵评分不需要）。
        is_rating: True 时走矩阵评分均值分支，False 时走矩阵单选频率分布分支。

    Returns:
        DataFrame；空时返回空 DataFrame。
    """

    # ------------------------------------------------------------------
    # 公共辅助：提取子题目标签
    # ------------------------------------------------------------------
    def _sub_label(col_title: str) -> str:
        for sep in ("—", "：", ":"):
            if sep in col_title:
                return col_title.rsplit(sep, 1)[-1].strip()
        return col_title

    def _safe_float(v: Any) -> Optional[float]:
        """将选项值转为浮点数；失败返回 None。"""
        try:
            return float(str(v))
        except (ValueError, TypeError):
            return None

    # ------------------------------------------------------------------
    # 分支一：矩阵评分 → 每子题两行：Mean + T2B%，列=分组
    # ------------------------------------------------------------------
    if is_rating:
        rows: List[dict] = []
        all_groups: List[str] = []

        for res in group_results:
            df_q = res["数据"]
            if df_q.empty:
                continue
            sub = _sub_label(str(res.get("题目", "")))
            groups_in_res = list(df_q["核心分组"].unique())

            # Mean 行
            mean_row: dict = {"子题目": sub, "指标": "Mean"}
            # T2B 行（4+5分占比）
            t2b_row: dict = {"子题目": "", "指标": "T2B%"}

            for grp in groups_in_res:
                grp_str = str(grp)
                if grp_str not in all_groups:
                    all_groups.append(grp_str)
                gdf = df_q[df_q["核心分组"] == grp]
                total_cnt = int(gdf["频次"].sum())
                if total_cnt == 0:
                    mean_row[grp_str] = ""
                    t2b_row[grp_str] = ""
                    continue
                weighted = sum(
                    fv * int(cnt)
                    for opt, cnt in zip(gdf["选项"], gdf["频次"])
                    if (fv := _safe_float(opt)) is not None
                )
                mean_row[grp_str] = round(weighted / total_cnt, 2)
                # T2B：4分+5分频次 / 该组有效人数
                t2b_cnt = sum(
                    int(cnt)
                    for opt, cnt in zip(gdf["选项"], gdf["频次"])
                    if (fv := _safe_float(opt)) is not None and fv >= 4
                )
                t2b_row[grp_str] = (t2b_cnt / total_cnt) if total_cnt > 0 else np.nan

            rows.append(mean_row)
            rows.append(t2b_row)

        if not rows:
            return pd.DataFrame()

        df_2d = pd.DataFrame(rows)
        ordered_cols = ["子题目", "指标"] + [g for g in all_groups if g in df_2d.columns]
        df_2d = df_2d[[c for c in ordered_cols if c in df_2d.columns]]
        df_2d = df_2d.fillna("")
        return df_2d

    # ------------------------------------------------------------------
    # 分支二：矩阵单选 → 每行=子题目，每列=量表选项（纯%），末尾加人数N
    # ------------------------------------------------------------------
    # 构建量表值→量表标签映射
    val_to_scale: Dict[str, str] = {}
    if scale_options:
        for label in scale_options:
            v = _extract_option_value(str(label))
            if v is not None:
                val_to_scale[str(int(v))] = label
                val_to_scale[f"{int(v)}.0"] = label
        if not val_to_scale:
            for pos, label in enumerate(scale_options):
                val_to_scale[str(pos + 1)] = label
                val_to_scale[f"{pos + 1}.0"] = label

    freq_rows: List[dict] = []
    n_per_sub: List[int] = []

    for res in group_results:
        df_q = res["数据"]
        if df_q.empty:
            continue
        sub = _sub_label(str(res.get("题目", "")))

        try:
            agg = df_q.groupby("选项")["频次"].sum().reset_index()
            row_total = int(agg["频次"].sum())
            row_data2: dict = {"子题目": sub}
            for _, r in agg.iterrows():
                pct = r["频次"] / row_total if row_total > 0 else 0
                col_key = val_to_scale.get(str(r["选项"]),
                          val_to_scale.get(f"{r['选项']}.0", str(r["选项"])))
                row_data2[col_key] = f"{pct:.1%}"
            freq_rows.append(row_data2)
            n_per_sub.append(row_total)
        except Exception:
            freq_rows.append({"子题目": sub})
            n_per_sub.append(0)

    if not freq_rows:
        return pd.DataFrame()

    df_2d = pd.DataFrame(freq_rows)

    # 按量表选项排序并补全缺失列（0.0%）
    if scale_options:
        for opt in scale_options:
            if opt not in df_2d.columns:
                df_2d[opt] = "0.0%"
        ordered = ["子题目"] + [o for o in scale_options if o in df_2d.columns]
        extra = [c for c in df_2d.columns if c not in ordered]
        df_2d = df_2d[[c for c in ordered + extra if c in df_2d.columns]]

    df_2d = df_2d.fillna("0.0%")

    # 末尾加人数N列
    df_2d["人数N"] = n_per_sub

    return df_2d


# ---------------------------------------------------------------------------
# Step 2：定量交叉分析
# ---------------------------------------------------------------------------
def _run_quant_cross(
    df: pd.DataFrame,
    column_type_map: Dict[str, str],
    segment_col_hint: Optional[str] = None,
    sig_test: bool = True,
    sig_alpha: float = 0.05,
) -> Tuple[List[dict], str, bool]:
    """确定分组列并执行全量交叉分析。

    Args:
        df: 问卷 DataFrame。
        column_type_map: 列名→题型映射。
        segment_col_hint: 来自 --segment-col 参数的子串（可为 None）。

    Returns:
        (cross_results, segment_col, is_synthetic)
    """
    print(f"\n[2/4] 定量交叉分析 ...")
    if sig_test and not (0 < sig_alpha < 1):
        print(f"  ⚠  --sig-alpha={sig_alpha} 超出有效范围 (0,1)，已回退为 0.05。")
        sig_alpha = 0.05

    df, segment_col, is_synthetic = _resolve_segment_col(df, column_type_map, segment_col_hint)

    n_groups = df[segment_col].nunique()
    print(
        f"  → 分组列：「{segment_col if not is_synthetic else '（总体）'}」"
        f"，共 {n_groups} 组：{list(df[segment_col].dropna().unique()[:5])}"
    )

    ignored_cols: set = {c for c, t in column_type_map.items() if t == "忽略"}
    ignored_cols.add(segment_col)

    question_types: Dict[str, List[int]] = {
        "单选": [], "多选": [], "评分": [], "矩阵单选": [], "矩阵评分": [],
    }
    explicit_single_cols: List[str] = []
    explicit_rating_cols: List[str] = []

    for col, t in column_type_map.items():
        if col in ignored_cols:
            continue
        if t == "单选":
            explicit_single_cols.append(col)
        elif t == "评分":
            explicit_rating_cols.append(col)
        elif t in ("矩阵单选", "矩阵评分", "矩阵"):
            # 矩阵子项按单选方式逐列分析（含无题号子项）
            # _export_results 再按 matrix_q_map 将它们重组为 2D 表
            # 使用 continue 跳过 question_types 累积，避免第一子项被 question_specs 重复处理
            explicit_single_cols.append(col)
            continue

        q_str = extract_qnum(str(col))
        if not q_str:
            continue
        try:
            q_num = int(q_str)
        except ValueError:
            continue

        if t == "单选":
            question_types["单选"].append(q_num)
        elif t == "多选":
            question_types["多选"].append(q_num)
        elif t == "评分":
            question_types["评分"].append(q_num)
        elif t in ("矩阵单选", "矩阵"):
            if q_num not in question_types["矩阵单选"] and q_num not in question_types["矩阵评分"]:
                question_types["矩阵单选"].append(q_num)
        elif t == "矩阵评分":
            if q_num not in question_types["矩阵单选"] and q_num not in question_types["矩阵评分"]:
                question_types["矩阵评分"].append(q_num)

    for k in question_types:
        question_types[k] = sorted(set(question_types[k]))

    type_counts = {k: len(v) for k, v in question_types.items() if v}
    print(f"  → 识别题型（含题号）：{type_counts}")
    print(
        f"  → 单选列（无题号）：{len(explicit_single_cols)} 列，"
        f"评分列（无题号）：{len(explicit_rating_cols)} 列"
    )

    question_specs = build_question_specs(df, question_types)
    results = run_quant_cross_engine(
        df,
        core_segment_col=segment_col,
        question_specs=question_specs,
        selected_cols_set=None,
        ignored_cols_set=ignored_cols,
        explicit_single_cols=explicit_single_cols,
        explicit_rating_cols=explicit_rating_cols,
        alpha=sig_alpha if sig_test else -1.0,
        min_group_size=5,
    )

    def _qnum_sort_key(res: dict) -> tuple:
        q = res.get("题目")
        q_str = extract_qnum(str(q)) if q else None
        q_num = int(q_str) if (q_str and q_str.isdigit()) else 9999
        return (q_num, _TYPE_RANK.get(str(res.get("题型")), 99), str(q))

    results = sorted(list(results or []), key=_qnum_sort_key)
    print(f"  → 完成，共 {len(results)} 道题的交叉分析结果。")
    return results, segment_col, is_synthetic


# ---------------------------------------------------------------------------
# Step 3：满意度回归建模
# ---------------------------------------------------------------------------
def _run_satisfaction_modeling(
    df: pd.DataFrame,
    column_type_map: Dict[str, str],
    n: int,
) -> Optional[dict]:
    """若满足条件则执行满意度多元回归，返回建模结果或 None。

    总样本量 n=len(df)：n<15 直接跳过；否则在成功返回的 dict 中附带
    ``is_low_sample_warning``（15<=n<50 为 True）与 ``low_sample_n``（=n）。

    Args:
        df: 问卷 DataFrame。
        column_type_map: 列名→题型映射。
        n: 样本量。

    Returns:
        regression_analysis 结果 dict（含 is_low_sample_warning、low_sample_n），或 None。
    """
    print(f"\n[3/4] 满意度建模 ...")

    if n < _MIN_N_REGRESSION_SKIP:
        print("  ⚠  样本量极小，无法进行回归计算")
        return None

    target_col: Optional[str] = None
    for col in df.columns:
        if col == _SYNTHETIC_SEGMENT_COL:
            continue
        lower = str(col).lower()
        if any(kw in lower for kw in _SATISFY_KEYWORDS):
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().mean() > 0.7:
                target_col = col
                break

    if target_col is None:
        print("  ⚠  未找到满意度/NPS类列（关键词：总体满意/满意度/NPS 等），跳过回归建模。")
        print("     提示：若数据中有满意度列，请确保列名包含上述关键词。")
        return None

    feature_cols = [
        col
        for col, t in column_type_map.items()
        if t == "评分" and col != target_col and col in df.columns
    ]

    if len(feature_cols) < 2:
        print(
            f"  ⚠  评分型特征列不足 2 个（找到 {len(feature_cols)} 个），跳过回归建模。"
            f"（目标列：{target_col}）"
        )
        return None

    print(f"  → 目标列：「{target_col}」，特征列数量：{len(feature_cols)}")

    try:
        from survey_tools.core.advanced_modeling import GameExperienceAnalyzer

        analyzer = GameExperienceAnalyzer(df)
        result = analyzer.regression_analysis(feature_cols, target_col)
        print(
            f"  → 回归建模完成，有效样本量：{result['sample_size']}，"
            f"Cronbach's α = {result['alpha']:.3f}"
        )
        result["is_low_sample_warning"] = _MIN_N_REGRESSION_SKIP <= n < _MIN_N_REGRESSION_FULL
        result["low_sample_n"] = n
        return result
    except Exception as e:
        print(f"  ⚠  回归建模异常：{e}")
        return None


# ---------------------------------------------------------------------------
# 占位：文本分析（待接入）
# ---------------------------------------------------------------------------
def _run_text_analysis_placeholder() -> None:
    # TODO: 后续接入 survey_tools/web/text_app.py 中的文本分析核心逻辑
    # 计划功能：关键词提取、主题聚类、情感分析、词频统计
    pass


# ---------------------------------------------------------------------------
# Step 4：汇总导出
# ---------------------------------------------------------------------------
def _export_results(
    df: pd.DataFrame,
    cross_results: List[dict],
    sig_test: bool,
    segment_col: str,
    is_synthetic: bool,
    regression_result: Optional[dict],
    output_dir: str,
    per_question_sheets: bool = False,
    outline: Optional[Dict[int, dict]] = None,
    matrix_q_map: Optional[Dict[str, int]] = None,
) -> Path:
    """将所有分析结果汇总为多 Sheet Excel，保存到 output_dir。

    Sheet 结构：
      - 样本概况：基础元数据
      - 交叉分析（汇总）：所有题目纵向拼接，始终输出
        - 矩阵题：合并为 2D 表（行=子题目，列=量表选项）
        - 单选/评分题：含均值行 + 有效填写人次行
      - [可选] 按 Q 题号分组的独立题 Sheet（--per-question-sheets 时输出）
      - 满意度回归结果（若有）：`is_low_sample_warning=True`（15<=N<50）时第 1 行为合并单元格免责（加粗红字），系数表浅灰底。

    Args:
        df: 原始问卷 DataFrame。
        cross_results: run_quant_cross_engine 返回的结果列表。
        sig_test: 是否开启显著性打标（星号/颜色）。
        segment_col: 实际使用的分组列名。
        is_synthetic: 是否为「总体」虚拟列。
        regression_result: regression_analysis 结果 dict 或 None。
        output_dir: 输出目录路径。
        per_question_sheets: 是否生成每题独立 Sheet。
        outline: _parse_outline_docx 解析结果（可为 None）。
        matrix_q_map: _infer_matrix_groups 返回的矩阵列→题号映射（可为 None）。

    Returns:
        最终写出文件的 Path 对象。
    """
    print(f"\n[4/4] 汇总导出 ...")

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    date_str = datetime.date.today().strftime("%Y%m%d")
    base_name = f"{date_str}_Playtest自动化分析报告"
    filename = output_path / f"{base_name}.xlsx"
    suffix_idx = 2
    while filename.exists():
        try:
            filename.open("a").close()
            break
        except PermissionError:
            filename = output_path / f"{base_name}_{suffix_idx}.xlsx"
            suffix_idx += 1

    sheets: List[Tuple[str, pd.DataFrame]] = []

    # Sheet 1：样本概况
    summary = pd.DataFrame(
        {
            "指标": [
                "总样本量", "总列数", "分析日期", "分组列", "分组方式",
                "交叉分析题目数", "是否有满意度回归结果", "是否加载问卷大纲",
            ],
            "值": [
                len(df),
                len(df.columns),
                datetime.date.today().isoformat(),
                segment_col if not is_synthetic else "（总体，无自然分组）",
                "全量频率分布（无分组对比）" if is_synthetic else "交叉分析（组间对比）",
                len(cross_results),
                "是" if regression_result else "否",
                "是" if outline else "否",
            ],
        }
    )
    sheets.append(("样本概况", summary))

    # 按 Q 题号分组（利用 matrix_q_map 修正矩阵子项的所属题号）
    q_group_map: Dict[str, List[dict]] = defaultdict(list)
    for res in cross_results:
        q = str(res.get("题目", ""))
        q_str = extract_qnum(q)
        if q_str and q_str.isdigit():
            key = f"Q{int(q_str):03d}"
        elif matrix_q_map and q in matrix_q_map:
            key = f"Q{matrix_q_map[q]:03d}"
        else:
            key = "其他"
        q_group_map[key].append(res)

    sorted_keys = sorted(
        q_group_map.keys(),
        key=lambda k: (int(k[1:]) if k != "其他" and k[1:].isdigit() else 99999, k),
    )

    # -----------------------------------------------------------------------
    # 辅助：将一个 DataFrame block 逐行写入 openpyxl worksheet
    # -----------------------------------------------------------------------
    def _safe_val(v: Any) -> Any:
        """写入友好的单元格值：空值转空字符串，其余尽量保留原类型。"""
        if v is None:
            return ""
        if isinstance(v, float) and pd.isna(v):
            return ""
        return v

    def _write_single_block(ws: Any, block_df: pd.DataFrame, start_row: int) -> int:
        """写单选/评分/多选题 block（block 已含题目标题、Banner N、列头、数据行），返回下一可用行号。"""
        cur = start_row
        for _, row in block_df.iterrows():
            for j, v in enumerate(row.values, 1):
                ws.cell(cur, j, _safe_val(v))
            cur += 1
        return cur

    def _write_matrix_block(
        ws: Any,
        matrix_2d: pd.DataFrame,
        title: str,
        start_row: int,
    ) -> int:
        """写矩阵题 2D block（题干行 + 列头行 + 数据行 + 空行），返回下一可用行号。"""
        # 题干行（仅写在第 1 列）
        ws.cell(start_row, 1, title)
        cur = start_row + 1
        # 列头行
        for j, h in enumerate(matrix_2d.columns, 1):
            ws.cell(cur, j, str(h))
        cur += 1
        # 数据行
        for _, row in matrix_2d.iterrows():
            for j, v in enumerate(row.values, 1):
                ws.cell(cur, j, _safe_val(v))
            cur += 1
        # 空行分隔
        cur += 1
        return cur

    # 视觉格式：色板与样式
    _FILL_TITLE = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    _FILL_HEADER = PatternFill(start_color="4A6785", end_color="4A6785", fill_type="solid")
    _FILL_ZEBRA = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _FILL_N_COL = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
    _FILL_SIG_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _FILL_SIG_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _FONT_TITLE = Font(bold=True, color="FFFFFF")
    _FONT_HEADER = Font(color="FFFFFF")
    _FONT_SIG_HIGH = Font(color="FF006100")
    _FONT_SIG_LOW = Font(color="FF9C0006")
    _ALIGN_RIGHT = Alignment(horizontal="right")

    def _parse_numeric(v: Any) -> Tuple[Optional[float], Optional[str]]:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if pd.isna(v):
                return None, None
            return float(v), "number"
        s = str(v).strip()
        if not s:
            return None, None
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0, "percent"
            except ValueError:
                return None, None
        try:
            return float(s), "number"
        except ValueError:
            return None, None

    def _mark_sig_cell(ws: Any, row: int, col: int, direction: str, metric: str, p_value: Optional[float] = None) -> None:
        cell = ws.cell(row, col)
        parsed, value_type = _parse_numeric(cell.value)
        if parsed is None:
            return
        cell.value = parsed
        star_lv = 2 if (p_value is not None and p_value < 0.01) else 1
        star_token = "**" if star_lv == 2 else "*"
        if metric in ("proportion", "mention_rate") or value_type == "percent":
            cell.number_format = f'0.0%"{star_token}"'
        else:
            cell.number_format = f'0.00"{star_token}"'
        if direction == "higher":
            cell.fill = _FILL_SIG_HIGH
            cell.font = _FONT_SIG_HIGH
        else:
            cell.fill = _FILL_SIG_LOW
            cell.font = _FONT_SIG_LOW

    def _apply_single_block_format(ws: Any, start_row: int, end_row: int, block_df: pd.DataFrame) -> None:
        """对单选/评分/多选题 block 应用视觉格式。"""
        n_cols = len(block_df.columns)
        # 确定人数N列索引（最后一列）
        n_col_idx = n_cols  # 1-based

        for r in range(start_row, end_row + 1):
            local_idx = r - start_row
            if local_idx >= len(block_df):
                break
            first_val = str(block_df.iloc[local_idx, 0])
            # 题目标题行
            if first_val.startswith("【"):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(r, c)
                    cell.fill = _FILL_TITLE
                    cell.font = _FONT_TITLE
            # Banner N / 列头行
            elif first_val == "样本量(N)" or (first_val == "选项" and local_idx > 0):
                for c in range(1, n_cols + 1):
                    ws.cell(r, c).fill = _FILL_HEADER
                    ws.cell(r, c).font = _FONT_HEADER
            # 本题平均分行
            elif "本题平均分" in first_val:
                for c in range(1, n_cols + 1):
                    ws.cell(r, c).fill = _FILL_HEADER
                    ws.cell(r, c).font = _FONT_HEADER
            # 数据行（含 T2B/B2B）
            elif first_val and first_val not in ("", "选项"):
                # 斑马纹：每 5 行一组，奇数组背景灰
                data_row_idx = local_idx - 4  # 约去标题、Banner、列头后的行号
                if data_row_idx >= 0 and (data_row_idx // 5) % 2 == 1:
                    for c in range(1, n_cols + 1):
                        ws.cell(r, c).fill = _FILL_ZEBRA
                # 人数N列背景
                for c in range(1, n_cols + 1):
                    cell = ws.cell(r, c)
                    if c == n_col_idx:
                        cell.fill = _FILL_N_COL
                    # 百分比列右对齐（非选项、非人数N 且含%）
                    elif c > 1 and "%" in str(cell.value):
                        cell.alignment = _ALIGN_RIGHT

    def _apply_matrix_block_format(
        ws: Any, start_row: int, end_row: int, matrix_2d: pd.DataFrame, has_rating: bool = False
    ) -> None:
        """对矩阵题 block 应用视觉格式。"""
        n_cols = len(matrix_2d.columns)
        # 题干行
        for c in range(1, n_cols + 1):
            cell = ws.cell(start_row, c)
            cell.fill = _FILL_TITLE
            cell.font = _FONT_TITLE
        # 列头行
        for c in range(1, n_cols + 1):
            cell = ws.cell(start_row + 1, c)
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
        # 数据行：斑马纹 + 右对齐数值
        for i in range(matrix_2d.shape[0]):
            row_num = start_row + 2 + i
            if (i // 5) % 2 == 1:
                for c in range(1, n_cols + 1):
                    ws.cell(row_num, c).fill = _FILL_ZEBRA
            for c in range(2, n_cols + 1):
                cell = ws.cell(row_num, c)
                if cell.value and str(cell.value).replace(".", "").replace("%", "").isdigit():
                    cell.alignment = _ALIGN_RIGHT
            # 矩阵评分的 T2B% 行：将比例显示为百分比
            if has_rating and "指标" in matrix_2d.columns:
                try:
                    metric_val = str(matrix_2d.iloc[i]["指标"])
                except Exception:
                    metric_val = ""
                if metric_val == "T2B%":
                    for c in range(3, n_cols + 1):
                        t_cell = ws.cell(row_num, c)
                        if isinstance(t_cell.value, (float, int)) and not isinstance(t_cell.value, bool):
                            t_cell.number_format = "0.0%"
        # 矩阵评分均值行：红黄绿三阶热力图（1→红，3→黄，5→绿）
        if has_rating and "指标" in matrix_2d.columns and n_cols >= 3:
            start_col_letter = get_column_letter(3)
            end_col_letter = get_column_letter(n_cols)
            for i in range(matrix_2d.shape[0]):
                try:
                    metric_val = str(matrix_2d.iloc[i]["指标"])
                except Exception:
                    metric_val = ""
                if metric_val != "Mean":
                    continue
                row_num = start_row + 2 + i
                ws.conditional_formatting.add(
                    f"{start_col_letter}{row_num}:{end_col_letter}{row_num}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=1,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=3,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=5,
                        end_color="63BE7B",
                    ),
                )

    # -----------------------------------------------------------------------
    # 构建各题 block 并收集 individual_sheets（供 --per-question-sheets）
    # -----------------------------------------------------------------------
    individual_sheets: List[Tuple[str, pd.DataFrame]] = []
    # 每个元素: ("single"|"matrix", block_df_or_matrix_2d, extra_info_dict)
    cross_blocks: List[Tuple[str, pd.DataFrame, dict]] = []
    # 记录需要打标的单元格坐标（row, col, direction, metric, p_value）
    sig_cells: List[Tuple[int, int, str, str, Optional[float]]] = []

    for q_key in sorted_keys:
        group_results = q_group_map[q_key]

        # 判断是否为矩阵题组（用 matrix_q_map 比依赖 res["题型"] 更可靠）
        is_matrix_group = matrix_q_map is not None and any(
            str(r.get("题目", "")) in matrix_q_map
            for r in group_results
        )

        # 获取大纲中该 Q 题的选项信息
        q_num_for_outline: Optional[int] = None
        if q_key != "其他" and q_key[1:].isdigit():
            q_num_for_outline = int(q_key[1:])
        outline_entry = outline.get(q_num_for_outline) if (outline and q_num_for_outline) else None

        if is_matrix_group:
            outline_q_type = (outline_entry or {}).get("type", "")
            # 矩阵量表题（评分矩阵）：腾讯问卷用"矩阵量表题"，问卷星用"矩阵评分"
            is_rating_matrix = any(t in outline_q_type for t in _RATING_OUTLINE_TYPES) or \
                                "矩阵评分" in outline_q_type
            # 矩阵单选不需要展开量表范围，矩阵评分也不需要 scale_options
            raw_scale = outline_entry["options"] if outline_entry else None
            scale_options = None if is_rating_matrix else _expand_scale_options(raw_scale)
            matrix_2d = _build_matrix_2d_table(
                group_results,
                scale_options=scale_options,
                is_rating=is_rating_matrix,
            )

            if not matrix_2d.empty:
                first_title = str(group_results[0].get("题目", ""))
                matrix_stem = (
                    first_title.rsplit(_MATRIX_SEP, 1)[0].strip()
                    if _MATRIX_SEP in first_title
                    else first_title
                )
                q_type_label = "矩阵评分" if is_rating_matrix else "矩阵单选"
                title_str = f"【{matrix_stem}】（{q_type_label}）"

                cross_blocks.append(("matrix", matrix_2d, {"title": title_str, "is_rating": is_rating_matrix}))

                # individual sheet：保留原有 pd.concat 格式
                def _make_row_2d(label: str, cols: List[str]) -> pd.DataFrame:
                    row = {c: "" for c in cols}
                    row[cols[0]] = label
                    return pd.DataFrame([row])
                ind_block = pd.concat(
                    [
                        _make_row_2d(title_str, list(matrix_2d.columns)),
                        matrix_2d,
                        _make_row_2d("", list(matrix_2d.columns)),
                    ],
                    ignore_index=True,
                )
                individual_sheets.append((
                    f"{q_key}_{matrix_stem[:20]}"[:31].translate(
                        str.maketrans(r"\/:*?[]", "       ")
                    ),
                    ind_block,
                ))
        else:
            group_parts: List[pd.DataFrame] = []
            for res in group_results:
                opt_list = outline_entry["options"] if outline_entry else None
                block = _build_question_block(res, option_list=opt_list)
                if not block.empty:
                    cross_blocks.append(("single", block, {"res": res}))
                    group_parts.append(block)

            if group_parts:
                group_df = pd.concat(group_parts, ignore_index=True)
                first_q = str(group_results[0].get("题目", ""))[:18]
                raw_name = f"{q_key}_{first_q}"
                for ch in r"\/:*?[]":
                    raw_name = raw_name.replace(ch, "")
                individual_sheets.append((raw_name[:31], group_df))

    # -----------------------------------------------------------------------
    # 写出 Excel（用 openpyxl 直写汇总 Sheet，避免 pd.concat 列对齐问题）
    # -----------------------------------------------------------------------
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Sheet 1：样本概况
        summary.to_excel(writer, sheet_name="样本概况", index=False)

        # Sheet 2：交叉分析（汇总）—— 逐题块从 A 列起写入
        wb = writer.book
        ws_cross = wb.create_sheet("交叉分析（汇总）")

        cur_row = 1
        if cross_blocks:
            for block_type, block_df, extra in cross_blocks:
                start_row = cur_row
                if block_type == "matrix":
                    cur_row = _write_matrix_block(
                        ws_cross, block_df, extra["title"], cur_row
                    )
                    end_row = cur_row - 2  # 不含空行
                    if end_row >= start_row:
                        _apply_matrix_block_format(
                            ws_cross, start_row, end_row, block_df,
                            has_rating=extra.get("is_rating", False),
                        )
                else:
                    cur_row = _write_single_block(ws_cross, block_df, cur_row)
                    end_row = cur_row - 2  # 不含空行
                    if end_row >= start_row:
                        _apply_single_block_format(
                            ws_cross, start_row, end_row, block_df
                        )
                        if sig_test:
                            res_obj = extra.get("res", {})
                            q_type = str(res_obj.get("题型", ""))
                            stats_obj = (res_obj or {}).get("stats", {}) or {}
                            ps = stats_obj.get("pipeline_summary", {}) or {}
                            # 评分题：标记均值行的分组列
                            if q_type == "评分":
                                direction_map = ps.get("direction_by_group", {}) or {}
                                mean_row_idx: Optional[int] = None
                                for i in range(len(block_df)):
                                    if str(block_df.iloc[i, 0]) == "本题平均分":
                                        mean_row_idx = i
                                        break
                                if mean_row_idx is not None:
                                    excel_row = start_row + mean_row_idx
                                    for col_idx, col_name in enumerate(block_df.columns, 1):
                                        info = direction_map.get(str(col_name))
                                        if not info:
                                            continue
                                        if info.get("is_significant"):
                                            sig_cells.append(
                                                (
                                                    excel_row,
                                                    col_idx,
                                                    str(info.get("direction", "higher")),
                                                    "mean",
                                                    float(info.get("p_value")) if info.get("p_value") is not None else None,
                                                )
                                            )
                            # 单/多选：按 option 行 + 分组列打标
                            for c in (ps.get("cells", []) or []):
                                if not c.get("is_significant"):
                                    continue
                                opt = str(c.get("option", ""))
                                grp = str(c.get("group", ""))
                                direction = str(c.get("direction", "higher"))
                                metric = str(c.get("metric", "proportion"))
                                row_idx: Optional[int] = None
                                for i in range(len(block_df)):
                                    if str(block_df.iloc[i, 0]) == opt:
                                        row_idx = i
                                        break
                                if row_idx is None:
                                    continue
                                col_idx: Optional[int] = None
                                for j, col_name in enumerate(block_df.columns, 1):
                                    if str(col_name) == grp:
                                        col_idx = j
                                        break
                                if col_idx is None:
                                    continue
                                p_val = c.get("p_value")
                                sig_cells.append(
                                    (
                                        start_row + row_idx,
                                        col_idx,
                                        direction,
                                        metric,
                                        float(p_val) if p_val is not None else None,
                                    )
                                )
            # 冻结第 1 列（选项/子题目列）
            ws_cross.freeze_panes = "B1"
            for rr, cc, direction, metric, p_value in sig_cells:
                _mark_sig_cell(ws_cross, rr, cc, direction, metric, p_value=p_value)
            # 底部图例说明
            legend_row = cur_row + 1
            ws_cross.cell(
                legend_row,
                1,
                "注：数值带 * 表示该群体在该项上具有统计学显著差异（* p<0.05，** p<0.01）。多选题显著性已进行 FDR(BH) 多重比较校正。",
            )
        else:
            ws_cross.cell(1, 1, "未执行交叉分析或无数据")

        # 可选：每题独立 Sheet
        if per_question_sheets:
            capped = individual_sheets[:_MAX_INDIVIDUAL_SHEETS]
            if len(individual_sheets) > _MAX_INDIVIDUAL_SHEETS:
                overflow = len(individual_sheets) - _MAX_INDIVIDUAL_SHEETS
                print(
                    f"  → 题目数（{len(individual_sheets)}）超过上限 {_MAX_INDIVIDUAL_SHEETS}，"
                    f"后 {overflow} 道题仅出现在「交叉分析（汇总）」Sheet 中。"
                )
            used_ind_names: set = set()
            for ind_name, ind_df in capped:
                base = ind_name[:28]
                safe = base
                sfx = 2
                while safe in used_ind_names:
                    safe = f"{base[:25]}_{sfx}"
                    sfx += 1
                used_ind_names.add(safe)
                ind_df.to_excel(writer, sheet_name=safe, index=False)

        # 满意度回归结果（低样本时 A1 免责声明 + 表区浅灰底）
        if regression_result:
            res_df = regression_result["results_df"]
            sheet_name = "满意度回归结果"
            low_warn = bool(regression_result.get("is_low_sample_warning"))
            low_n = int(regression_result.get("low_sample_n", len(df)))
            if low_warn:
                res_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                ws_reg = writer.sheets[sheet_name]
                n_cols = int(res_df.shape[1])
                end_letter = get_column_letter(n_cols)
                ws_reg.merge_cells(f"A1:{end_letter}1")
                txt = _SATISFACTION_LOW_N_DISCLAIMER.format(n=low_n)
                c1 = ws_reg["A1"]
                c1.value = txt
                c1.font = _FONT_SATISFACTION_WARN
                c1.alignment = Alignment(wrap_text=True, vertical="center")
                # 表头+数据行（pandas startrow=1 → 表头在第 2 行）
                for r in range(2, 2 + len(res_df) + 1):
                    for cc in range(1, n_cols + 1):
                        ws_reg.cell(r, cc).fill = _FILL_SATISFACTION_LOW_N
            else:
                res_df.to_excel(writer, sheet_name=sheet_name, index=False)

    n_sheets = 2 + (len(individual_sheets) if per_question_sheets else 0) + (1 if regression_result else 0)
    n_individual = len(individual_sheets) if per_question_sheets else 0
    detail = f"，含 {n_individual} 道题独立 Sheet" if per_question_sheets and n_individual > 0 else ""
    print(f"  → 报告已保存：{filename.resolve()}")
    print(
        f"  → 共写出 {n_sheets} 个 Sheet"
        f"（样本概况 + 汇总{detail}{' + 回归结果' if regression_result else ''}）。"
    )
    return filename


# ---------------------------------------------------------------------------
# 公共入口：支持外部传入 DataFrame（Web 端复用）
# ---------------------------------------------------------------------------
def _dispatch_parse_outline(p: Path) -> Dict[int, dict]:
    """根据扩展名分派到对应大纲解析器。"""
    if p.suffix.lower() == ".txt":
        return parse_outline_txt(p)
    return parse_outline_docx(p)


def load_parsed_outline(
    data_dir: str,
    outline: Optional[str] = None,
) -> Optional[Dict[int, dict]]:
    """加载问卷大纲（优先显式路径，其次自动发现）。"""
    parsed_outline: Optional[Dict[int, dict]] = None
    try:
        if outline:
            outline_path = Path(outline)
            if outline_path.exists():
                fmt = "腾讯问卷 .txt" if outline_path.suffix.lower() == ".txt" else "问卷星 .docx"
                print(f"\n  → [--outline] 已指定大纲文件：{outline_path.name}（{fmt}）")
                parsed_outline = _dispatch_parse_outline(outline_path)
                print(f"  → 大纲解析完成，共 {len(parsed_outline)} 道题。")
            else:
                print(f"\n  ⚠  --outline 指定的文件不存在：{outline_path}，忽略。")
        else:
            auto_outline = _get_latest_local_outline(data_dir)
            if auto_outline:
                fmt = "腾讯问卷 .txt" if auto_outline.suffix.lower() == ".txt" else "问卷星 .docx"
                print(f"\n  → 自动发现大纲文件：{auto_outline.name}（{fmt}）")
                parsed_outline = _dispatch_parse_outline(auto_outline)
                print(f"  → 大纲解析完成，共 {len(parsed_outline)} 道题。")
            else:
                print(f"\n  → 未在 {data_dir} 中发现大纲文件（.docx / .txt），题型和选项将依赖自动识别。")
    except Exception as _e:
        print(f"\n  ⚠  大纲解析失败（已跳过）：{_e}")
        parsed_outline = None
    return parsed_outline


def run_pipeline(
    df: pd.DataFrame,
    output_dir: str,
    segment_col: Optional[str] = None,
    per_question_sheets: bool = False,
    outline: Optional[Dict[int, dict]] = None,
    sig_test: bool = True,
    sig_alpha: float = 0.05,
    sav_meta: Optional[Any] = None,
) -> Dict[str, Any]:
    """执行 Playtest 流水线核心调度（可被 CLI / Web 复用）。"""
    n = len(df)
    matrix_q_map = _infer_matrix_groups(df.columns.tolist())
    if matrix_q_map:
        print(
            f"  → 识别到矩阵题子项列：{len(matrix_q_map)} 列，"
            f"涉及 {len(set(matrix_q_map.values()))} 道矩阵题。"
        )

    print(f"\n  正在自动识别题型（共 {len(df.columns)} 列）...")
    column_type_map = _auto_classify_columns(
        df,
        sav_meta=sav_meta,
        outline=outline,
        matrix_q_map=matrix_q_map,
    )
    type_summary: Dict[str, int] = defaultdict(int)
    for t in column_type_map.values():
        type_summary[t] += 1
    print(f"  题型统计：{dict(type_summary)}")

    cross_results: List[dict] = []
    resolved_segment_col: str = _SYNTHETIC_SEGMENT_COL
    is_synthetic = True
    try:
        cross_results, resolved_segment_col, is_synthetic = _run_quant_cross(
            df,
            column_type_map,
            segment_col_hint=segment_col,
            sig_test=sig_test,
            sig_alpha=sig_alpha,
        )
        if not sig_test:
            print("  → 显著性检验已关闭（--no-sig-test），将跳过显著性打标与颜色标注。")
    except Exception as e:
        print(f"\n  ✗ 交叉分析失败：{e}")

    regression_result: Optional[dict] = None
    try:
        regression_result = _run_satisfaction_modeling(df, column_type_map, n)
    except Exception as e:
        print(f"\n  ✗ 满意度建模失败：{e}")

    _run_text_analysis_placeholder()

    output_file = _export_results(
        df,
        cross_results,
        sig_test,
        resolved_segment_col,
        is_synthetic,
        regression_result,
        output_dir,
        per_question_sheets=per_question_sheets,
        outline=outline,
        matrix_q_map=matrix_q_map,
    )
    return {
        "output_file": output_file,
        "cross_results": cross_results,
        "resolved_segment_col": resolved_segment_col,
        "is_synthetic": is_synthetic,
        "regression_result": regression_result,
        "column_type_map": column_type_map,
        "matrix_q_map": matrix_q_map,
    }


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------
@click.command()
@click.option(
    "--data-dir",
    default="data/raw",
    show_default=True,
    help="存放原始问卷数据的目录（.sav/.csv/.xlsx）。",
)
@click.option(
    "--output-dir",
    default="data/processed",
    show_default=True,
    help="分析报告的输出目录。",
)
@click.option(
    "--segment-col",
    default=None,
    help="[C] 指定分组列（支持子串模糊匹配，如 '玩家类型'）。缺省时自动识别；"
    "完全找不到时自动用「总体」代替。",
)
@click.option(
    "--per-question-sheets",
    is_flag=True,
    default=False,
    help="[D] 在汇总 Sheet 之外，额外为每道题生成独立 Sheet（最多 60 个）。"
    "默认关闭，报告只有「样本概况」「汇总」「回归结果」3 个 Sheet。",
)
@click.option(
    "--outline",
    default=None,
    help="[大纲] 指定问卷大纲文件路径（.docx 问卷星格式 或 .txt 腾讯问卷格式）。"
    "若不指定，自动发现 data-dir 下 mtime 最新的 .docx 或 .txt 文件。"
    "大纲用于修正题型识别、补全 0% 选项、规范选项排序。",
)
@click.option(
    "--sig-test/--no-sig-test",
    default=True,
    show_default=True,
    help="是否启用组间均值显著性检验（独立样本T检验）及导出红字标记。",
)
@click.option(
    "--sig-alpha",
    default=0.05,
    show_default=True,
    type=float,
    help="显著性阈值 alpha（满足 p-value < alpha 视为显著）。",
)
@click.option(
    "--sheet-name",
    default=None,
    help="Excel 读取的 Sheet（支持索引如 0，或名称如 Sheet1）。仅对 .xlsx/.xls 生效。",
)
def main(
    data_dir: str,
    output_dir: str,
    segment_col: Optional[str],
    per_question_sheets: bool,
    outline: Optional[str],
    sig_test: bool,
    sig_alpha: float,
    sheet_name: Optional[str],
) -> None:
    """Playtest 自动化分析流水线 v0.3。

    自动读取最新问卷数据 → 识别题型 → 交叉/频率分析 → 满意度建模 → 导出多 Sheet Excel。
    支持问卷大纲导入（.docx 问卷星格式 / .txt 腾讯问卷格式），提升矩阵题识别精度与选项完整性。
    """
    print("=" * 60)
    print("  Playtest 自动化分析流水线  v0.3")
    print("=" * 60)

    parsed_outline = load_parsed_outline(data_dir=data_dir, outline=outline)

    # Step 1：数据装载
    parsed_sheet_name: Optional[int | str] = 0
    if sheet_name is not None and str(sheet_name).strip() != "":
        raw_sheet = str(sheet_name).strip()
        parsed_sheet_name = int(raw_sheet) if re.fullmatch(r"-?\d+", raw_sheet) else raw_sheet
    try:
        df, sav_meta = _load_data(data_dir, sheet_name=parsed_sheet_name)
    except FileNotFoundError as e:
        print(f"\n✗ {e}")
        sys.exit(1)
    try:
        run_res = run_pipeline(
            df=df,
            output_dir=output_dir,
            segment_col=segment_col,
            per_question_sheets=per_question_sheets,
            outline=parsed_outline,
            sig_test=sig_test,
            sig_alpha=sig_alpha,
            sav_meta=sav_meta,
        )
    except Exception as e:
        print(f"\n  ✗ 导出失败：{e}")
        sys.exit(1)

    output_file = run_res["output_file"]
    print("\n" + "=" * 60)
    print("  ✅ 流水线执行完毕！")
    print(f"  报告路径：{output_file.resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
